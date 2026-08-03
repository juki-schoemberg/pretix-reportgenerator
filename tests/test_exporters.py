"""Tests for ``pretix_custom_reports/exporters.py``.

Owner from wave 2 on: exporter-dev (ORCHESTRIERUNG.md section 5).

The module under test is thin on purpose -- rows come from the query compiler,
bytes come from ``ListExporter`` -- so these tests concentrate on the seams
where a saved report meets pretix' export machinery, and on the four failure
modes that would otherwise be invisible:

1. a scheduled export whose report was deleted (five Celery retries and the
   word "Internal Error" instead of a sentence naming the report),
2. a multi-event export over an event where a column does not resolve,
3. a relative date filter that gets frozen at save time instead of being
   re-evaluated per run,
4. CSV/XLSX formula injection -- which pretix already neutralises, so the test
   asserts that it happens rather than adding a second layer.

Registration comes from ``signals.py`` (integrator, wave 4), which connects the
two receivers of this module at plugin import time -- the copy-ready lines from
handoff/requests/exporter-dev-an-integrator-signals.md. The ``registered``
fixture no longer establishes that wiring, it only *guarantees* it and asserts
that the functions behind the two dispatch_uids are still ours, so these tests
fail the moment the wiring stops matching reality. See the fixture for why it
must not simply connect and disconnect around every test.
"""

import datetime
import pytest
import warnings
import weakref
from decimal import Decimal
from django.core import mail as djmail
from django.utils.timezone import now
from django_scopes import scope, scopes_disabled
from freezegun import freeze_time
from pretix.base.exporter import ListExporter
from pretix.base.models import (
    Event,
    Item,
    Order,
    OrderPosition,
    Question,
    QuestionAnswer,
    ScheduledEventExport,
    ScheduledOrganizerExport,
    Team,
    User,
)
from pretix.base.services.export import (
    ExportError,
    init_event_exporter,
    init_event_exporters,
    init_organizer_exporters,
    run_scheduled_exports,
)
from pretix.base.signals import (
    register_data_exporters,
    register_multievent_data_exporters,
)
from zoneinfo import ZoneInfo

from pretix_custom_reports import contracts, exporters
from pretix_custom_reports.contracts import (
    BooleanStyle,
    ColumnFormat,
    DataType,
    DateStyle,
    NumberStyle,
)
from pretix_custom_reports.models import ReportDefinition

DISPATCH_UID = "pretix_custom_reports_exporter"
MULTI_DISPATCH_UID = "pretix_custom_reports_multiexporter"

#: The wiring the tests need, as ``(signal, receiver, dispatch_uid)``. Same
#: three tuples ``signals.py`` uses, which is the point.
WIRING = (
    (register_data_exporters, exporters.register_report_exporter, DISPATCH_UID),
    (
        register_multievent_data_exporters,
        exporters.register_multievent_report_exporter,
        MULTI_DISPATCH_UID,
    ),
)


def connected_receiver(signal, dispatch_uid):
    """The receiver currently connected to *signal* under *dispatch_uid*.

    ``Signal.receivers`` holds ``(lookup_key, receiver, is_async)`` in Django
    5.2 and ``lookup_key`` is ``(dispatch_uid or id(receiver), id(sender))``
    (django/dispatch/dispatcher.py:96-99, 113-117). With the default
    ``weak=True`` the second slot is a ``weakref.ref``, so it has to be
    dereferenced before it can be compared to a function.

    Returns ``None`` if nothing is connected under that dispatch_uid.
    """
    for lookup_key, receiver, *_ in signal.receivers:
        if lookup_key[0] == dispatch_uid:
            if isinstance(receiver, weakref.ReferenceType):
                return receiver()
            return receiver
    return None


#: Snapshot of the production wiring, taken while this module is imported and
#: therefore before any test or fixture in it has run. ``apps.ready()`` imports
#: ``signals.py`` exactly once per process, so this is what the plugin
#: established at startup -- and the state every test in this file has to hand
#: back untouched.
WIRING_AT_IMPORT = {
    dispatch_uid: connected_receiver(signal, dispatch_uid)
    for signal, _receiver, dispatch_uid in WIRING
}


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module", autouse=True)
def wiring_before_this_module():
    """The signal wiring as this module found it, for the canary at the bottom.

    Module scoped and autouse, so it is established before the first test here
    runs. Deliberately *not* the same thing as :data:`WIRING_AT_IMPORT`: what
    this file has to answer for is that it changes nothing, not that the
    session was healthy when it got the process. Somebody else's leak is
    somebody else's test to fail.
    """
    return {
        dispatch_uid: connected_receiver(signal, dispatch_uid)
        for signal, _receiver, dispatch_uid in WIRING
    }


@pytest.fixture
def registered():
    """Guarantee the two receivers are connected -- and restore the state after.

    Since wave 4 ``signals.py`` connects both receivers at plugin import
    (``apps.ready()``), so in a normal run there is nothing left to do here.
    This fixture must therefore not be written as a plain connect/disconnect
    pair, and that is not a style preference:

    ``Signal.connect()`` skips a receiver whose ``(dispatch_uid, sender_id)``
    key is already present (dispatcher.py:113-117), and
    ``Signal.disconnect(dispatch_uid=...)`` matches on that key *alone* --
    the receiver argument is ignored entirely (dispatcher.py:138-153). Neither
    call knows or cares who connected first. A fixture that connected and then
    disconnected ``DISPATCH_UID`` would hand back a process in which the
    *production* registration is gone, for every test that follows in the same
    pytest session, in whatever file. ``signals.py`` runs once and does not
    re-connect itself. pretix' ``EventPluginSignal``/``OrganizerPluginSignal``
    override ``connect`` but not ``disconnect`` (pretix/base/signals.py:261-311),
    so none of this is softened for plugin signals.

    So: connect only what is missing, disconnect only what we connected. If the
    dispatch_uid is already taken, assert it is taken by the function we expect
    -- otherwise the no-op ``connect()`` above would quietly run the whole
    module against somebody else's receiver.

    ``register_multievent_data_exporters`` is an
    ``OrganizerPluginSignal(allow_legacy_plugins=True)`` and this plugin is
    event level, so ``connect()`` emits a ``DeprecationWarning``
    (pretix/base/signals.py:301-306). pretix' own test config filters it; ours
    does not, so it is silenced here -- and only here, deliberately narrow, so
    that a *different* deprecation would still be visible.
    """
    connected_by_us = []
    for signal, receiver, dispatch_uid in WIRING:
        existing = connected_receiver(signal, dispatch_uid)
        if existing is not None:
            assert existing is receiver, (
                f"{dispatch_uid!r} is connected to {existing!r}, not to "
                f"{receiver!r} -- signals.py and exporters.py disagree."
            )
            continue
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message=".*organizer-level.*",
                category=DeprecationWarning,
            )
            signal.connect(receiver, dispatch_uid=dispatch_uid)
        connected_by_us.append((signal, dispatch_uid))
    yield
    for signal, dispatch_uid in connected_by_us:
        signal.disconnect(dispatch_uid=dispatch_uid)


@pytest.fixture
def orders(event):
    """Two paid orders, one canceled position, one test mode order.

    The test mode order and the canceled position exist so that the run-time
    overrides have something to switch on and off.
    """
    with scopes_disabled():
        channel = event.organizer.sales_channels.get(identifier="web")
        item = Item.objects.create(
            event=event, name="Ticket", internal_name="ticket", default_price=23
        )
        first = Order.objects.create(
            event=event,
            code="AAAAA",
            status=Order.STATUS_PAID,
            email="a@example.org",
            sales_channel=channel,
            datetime=now() - datetime.timedelta(days=2),
            expires=now() + datetime.timedelta(days=10),
            total=Decimal("23.00"),
            comment="plain",
        )
        OrderPosition.objects.create(
            order=first, item=item, price=Decimal("23.00"), positionid=1
        )
        OrderPosition.all.create(
            order=first,
            item=item,
            price=Decimal("5.00"),
            positionid=2,
            canceled=True,
        )
        second = Order.objects.create(
            event=event,
            code="BBBBB",
            status=Order.STATUS_PENDING,
            email="b@example.org",
            sales_channel=channel,
            datetime=now() - datetime.timedelta(days=40),
            expires=now() + datetime.timedelta(days=10),
            total=Decimal("11.00"),
            comment="plain",
        )
        OrderPosition.objects.create(
            order=second, item=item, price=Decimal("11.00"), positionid=1
        )
        testmode = Order.objects.create(
            event=event,
            code="CCCCC",
            status=Order.STATUS_PAID,
            email="c@example.org",
            sales_channel=channel,
            testmode=True,
            datetime=now() - datetime.timedelta(days=1),
            expires=now() + datetime.timedelta(days=10),
            total=Decimal("7.00"),
            comment="plain",
        )
        OrderPosition.objects.create(
            order=testmode, item=item, price=Decimal("7.00"), positionid=1
        )
        return {"item": item, "first": first, "second": second, "testmode": testmode}


def make_report(event=None, organizer=None, identifier="codes", **kwargs):
    """A minimal saved report: one column, ``order.code``."""
    definition = {
        "schema_version": contracts.SCHEMA_VERSION,
        "base": "order",
        "columns": [{"field": "order.code"}],
    }
    definition.update(kwargs.pop("definition", {}))
    with scopes_disabled():
        return ReportDefinition.objects.create(
            event=event,
            organizer=organizer,
            name=kwargs.pop("name", "Order codes"),
            identifier=identifier,
            definition=definition,
            **kwargs,
        )


def exporter_for_event(event, user, **kwargs):
    """The exporter instance pretix itself would build for an event export."""
    return init_event_exporter(
        identifier=exporters.CustomReportExporter.identifier,
        event=event,
        user=user,
        **kwargs,
    )


def exporter_for_organizer(organizer, user, **kwargs):
    for ex in init_organizer_exporters(organizer=organizer, user=user, **kwargs):
        if ex.identifier == exporters.CustomReportExporter.identifier:
            return ex
    return None


def csv_form_data(identifier="codes", **extra):
    data = {"_format": "default", contracts.EXPORT_FORM_REPORT_KEY: identifier}
    data.update(extra)
    return data


def rows_of(exporter, form_data):
    """Everything ``iterate_list`` yields, minus the progress marker."""
    return [
        line
        for line in exporter.iterate_list(form_data)
        if not isinstance(line, ListExporter.ProgressSetTotal)
    ]


# ---------------------------------------------------------------------------
# 1. It shows up where pretix looks for exporters
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_exporter_appears_in_the_event_export_ui(registered, event, user_with_perms):
    """``init_event_exporters`` is what the event export page enumerates."""
    with scope(organizer=event.organizer):
        found = [
            ex.identifier for ex in init_event_exporters(event, user=user_with_perms)
        ]
    assert exporters.CustomReportExporter.identifier in found


@pytest.mark.django_db
def test_exporter_appears_in_the_organizer_export_ui(
    registered, event, user_with_perms
):
    """``init_organizer_exporters`` is what the organizer export page enumerates."""
    with scope(organizer=event.organizer):
        ex = exporter_for_organizer(event.organizer, user_with_perms)
    assert ex is not None
    assert ex.is_multievent is True
    # The multi-event instance is handed a queryset, not an event.
    assert ex.event is None


@pytest.mark.django_db
def test_exporter_is_hidden_from_a_user_without_the_orders_permission(
    registered, event, user_without_perms
):
    """We keep ``ListExporter``'s ``event.orders:read`` requirement.

    ``init_event_exporters`` drops every exporter the acting account lacks the
    permission for (services/export.py:211-213). This asserts that we did not
    weaken that by overriding ``get_required_event_permission``.
    """
    assert (
        exporters.CustomReportExporter.get_required_event_permission()
        == "event.orders:read"
    )
    with scope(organizer=event.organizer):
        found = [
            ex.identifier for ex in init_event_exporters(event, user=user_without_perms)
        ]
    assert exporters.CustomReportExporter.identifier not in found


@pytest.mark.django_db
def test_the_format_choice_of_listexporter_survives(registered, event, user_with_perms):
    """Overriding ``export_form_fields`` would delete ``_format`` and break render.

    docs/pretix-api-notes.md section 1, pitfall 4. The check is cheap and the
    failure mode (an export that silently produces nothing) is not.
    """
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        fields = ex.export_form_fields
    assert "_format" in fields
    assert list(fields)[0] == "_format"
    assert set(dict(fields["_format"].choices)) == set(exporters.EXPORT_FORMATS)
    assert contracts.EXPORT_FORM_REPORT_KEY in fields


@pytest.mark.django_db
def test_the_report_choices_come_from_the_event(registered, event, user_with_perms):
    with scopes_disabled():
        make_report(event=event, identifier="codes", name="Order codes")
        make_report(event=event, identifier="other", name="Second report")
        other_event = Event.objects.create(
            organizer=event.organizer,
            name="Other",
            slug="other",
            date_from=now(),
            plugins="pretix_custom_reports",
        )
        make_report(event=other_event, identifier="foreign", name="Not mine")
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        choices = dict(ex.report_choices())
    assert set(choices) == {"codes", "other"}


@pytest.mark.django_db
def test_multievent_report_choices_are_collapsed_by_identifier(
    registered, event, user_with_perms
):
    """The same identifier in two events is one choice, not two.

    That is the whole reason the reference is an identifier and not a primary
    key (ADR 0001 section 5.1): an event copy carries the identifier along, and
    a multi-event export has to resolve it once per event.
    """
    with scopes_disabled():
        second = Event.objects.create(
            organizer=event.organizer,
            name="Second",
            slug="second",
            date_from=now(),
            plugins="pretix_custom_reports",
        )
        make_report(event=event, identifier="codes", name="Order codes")
        make_report(event=second, identifier="codes", name="Order codes (copy)")
    with scope(organizer=event.organizer):
        ex = exporter_for_organizer(event.organizer, user_with_perms)
        choices = dict(ex.report_choices())
    assert list(choices) == ["codes"]
    assert "codes" in choices["codes"]


@pytest.mark.django_db
def test_report_choices_never_leave_the_events_we_were_given(
    registered, event, user_with_perms
):
    """The permission boundary in one assertion.

    ``self.events`` is what pretix restricted to the acting account. A report of
    an event outside that queryset must not be offered even though it lives in
    the same organizer.
    """
    with scopes_disabled():
        hidden = Event.objects.create(
            organizer=event.organizer,
            name="Hidden",
            slug="hidden",
            date_from=now(),
            plugins="pretix_custom_reports",
        )
        make_report(event=event, identifier="visible")
        make_report(event=hidden, identifier="invisible")
    with scope(organizer=event.organizer):
        ex = exporter_for_organizer(
            event.organizer,
            user_with_perms,
            event_qs=Event.objects.filter(pk=event.pk),
        )
        choices = dict(ex.report_choices())
    assert set(choices) == {"visible"}


# ---------------------------------------------------------------------------
# 2. A report runs
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_a_report_runs_end_to_end_as_csv(registered, event, user_with_perms, orders):
    with scopes_disabled():
        make_report(event=event, identifier="codes")
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        filename, mimetype, content = ex.render(csv_form_data())
    body = content.decode("utf-8")
    assert mimetype == "text/csv"
    assert filename.endswith(".csv")
    assert "codes" in filename
    # Header row plus the two non-testmode orders.
    assert body.splitlines()[0].strip('"') == "Order code"
    assert "AAAAA" in body and "BBBBB" in body
    assert "CCCCC" not in body  # test mode is excluded by default


def render_xlsx(exporter, form_data):
    """Render to a buffer instead of letting pretix buffer it for us.

    ``ListExporter._render_xlsx`` without ``output_file`` writes to a
    ``tempfile.NamedTemporaryFile`` and re-opens it by name
    (pretix/base/exporter.py:322-326). On Windows a NamedTemporaryFile cannot be
    re-opened while it is still open, so that branch raises ``PermissionError``
    there. That is an upstream platform limitation, not something this plugin
    can fix or should work around in production code -- and the ``output_file``
    branch runs the identical ``SafeWorkbook`` code, so the tests use it and
    stay honest on every platform.
    """
    import io

    buffer = io.BytesIO()
    filename, mimetype, content = exporter.render(form_data, output_file=buffer)
    assert content is None  # documented: with output_file, bytes are not returned
    return filename, mimetype, buffer.getvalue()


@pytest.mark.django_db
def test_a_report_runs_end_to_end_as_xlsx(registered, event, user_with_perms, orders):
    with scopes_disabled():
        make_report(event=event, identifier="codes")
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        filename, mimetype, content = render_xlsx(
            ex, {"_format": "xlsx", contracts.EXPORT_FORM_REPORT_KEY: "codes"}
        )
    assert filename.endswith(".xlsx")
    assert mimetype.endswith("spreadsheetml.sheet")
    assert content[:2] == b"PK"  # a zip container, i.e. a real xlsx


@pytest.mark.django_db
def test_running_a_report_writes_an_execution_log_entry(
    registered, event, user_with_perms, orders
):
    """``log_executed`` had no caller before this module (handoff/status/persistence-dev.md)."""
    with scopes_disabled():
        report = make_report(event=event, identifier="codes")
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        ex.render(csv_form_data())
    with scopes_disabled():
        entries = list(report.all_logentries())
    assert len(entries) == 1
    assert entries[0].action_type == contracts.LOG_ACTION_EXECUTED
    assert entries[0].user == user_with_perms
    assert entries[0].parsed_data["row_count"] == 2
    assert entries[0].parsed_data["format"] == "default"


@pytest.mark.django_db
def test_runtime_overrides_change_the_result_without_touching_the_saved_report(
    registered, event, user_with_perms, orders
):
    """Overrides are per run. The stored definition must come out unchanged."""
    with scopes_disabled():
        report = make_report(event=event, identifier="codes")
        stored_before = dict(report.definition)
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        rows = rows_of(
            ex,
            csv_form_data(
                **{
                    exporters.FORM_KEY_INCLUDE_TESTMODE: exporters.OVERRIDE_YES,
                    exporters.FORM_KEY_ROW_LIMIT: 2,
                }
            ),
        )
    codes = [row[0] for row in rows[1:]]
    assert len(codes) == 2  # row_limit
    with scopes_disabled():
        report.refresh_from_db()
    assert report.definition == stored_before


@pytest.mark.django_db
def test_the_row_limit_override_is_range_checked(
    registered, event, user_with_perms, orders
):
    """``export_form_data`` is never revalidated, so we do it (api-notes 5, pitfall 1)."""
    with scopes_disabled():
        make_report(event=event, identifier="codes")
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        for bad in ("many", 0, -1, contracts.MAX_ROW_LIMIT + 1, True, [3]):
            with pytest.raises(ExportError) as excinfo:
                ex.render(csv_form_data(**{exporters.FORM_KEY_ROW_LIMIT: bad}))
            assert "row limit" in str(excinfo.value)


@pytest.mark.django_db
def test_an_unknown_format_says_so_instead_of_pretending_the_export_was_empty(
    registered, event, user_with_perms, orders
):
    """``ListExporter.render`` has no ``else`` branch (api-notes 1, pitfall 1).

    Without our check, ``render`` returns ``None`` and the calling task reports
    "Your export did not contain any data." -- which sends the owner looking for
    missing orders instead of a broken configuration.
    """
    with scopes_disabled():
        make_report(event=event, identifier="codes")
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        assert ListExporter.render(ex, {"_format": "ods"}) is None
        with pytest.raises(ExportError) as excinfo:
            ex.render({"_format": "ods", contracts.EXPORT_FORM_REPORT_KEY: "codes"})
    assert "ods" in str(excinfo.value)


@pytest.mark.django_db
def test_a_missing_or_malformed_report_reference_is_an_exporterror(
    registered, event, user_with_perms, orders
):
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        for bad in (None, "", 17, {"pk": 1}, "not a valid identifier!"):
            with pytest.raises(ExportError):
                ex.render({"_format": "default", contracts.EXPORT_FORM_REPORT_KEY: bad})


# ---------------------------------------------------------------------------
# 2b. Column formats reach the file (finding T-001)
# ---------------------------------------------------------------------------
#
# The bug these tests were written for: ``date_style`` was offered by the
# editor, honoured by the preview and dropped by the export, so "date only"
# produced a full timestamp in the file. Two styles, one output line
# (handoff/blockers.md, T-001).
#
# What is asserted here is the *pipeline*: a style saved in a definition changes
# the bytes. The parity between this pipeline and the preview is asserted
# separately, in test_the_preview_and_the_export_share_one_renderer.


def make_format_report(event, columns, identifier="fmt"):
    """A report over exactly one order (``AAAAA``) with the given columns.

    Filtered down to a single row on purpose: these tests compare cells, and a
    second row would only add ordering questions that have nothing to do with
    formatting.
    """
    return make_report(
        event=event,
        identifier=identifier,
        name="Formats",
        definition={
            "schema_version": contracts.SCHEMA_VERSION,
            "base": "order",
            "columns": columns,
            "filters": {
                "op": "and",
                "children": [
                    {"field": "order.code", "operator": "exact", "value": "AAAAA"}
                ],
            },
        },
    )


def first_data_row(event, user_with_perms, columns):
    """The one data row of a report with *columns*, as the exporter yields it.

    Cell values before ``ListExporter`` serialises them, which is where the
    native/formatted distinction is still visible.
    """
    with scopes_disabled():
        ReportDefinition.objects.filter(event=event, identifier="fmt").delete()
        make_format_report(event, columns)
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        rows = rows_of(ex, csv_form_data(identifier="fmt"))
    assert len(rows) == 2, rows  # header plus the single order
    return rows[1]


def csv_data_line(event, user_with_perms, columns):
    """The one data line of a report with *columns*, as bytes-turned-text."""
    with scopes_disabled():
        ReportDefinition.objects.filter(event=event, identifier="fmt").delete()
        make_format_report(event, columns)
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        _fn, _mime, content = ex.render(csv_form_data(identifier="fmt"))
    return content.decode("utf-8").splitlines()[1]


@pytest.mark.django_db
def test_a_date_style_chosen_in_the_editor_reaches_the_csv(
    registered, event, user_with_perms, orders
):
    """The reproduction from the blocker, as an assertion instead of an xfail.

    ``iso`` and ``date_only`` used to produce the identical line, both carrying
    the full timestamp. They now differ, and they differ in the direction the
    user asked for: no time of day in the date-only file.
    """
    columns = [{"field": "order.code"}, {"field": "order.datetime"}]
    lines = {}
    for style in ("iso", "date_only"):
        columns[1]["format"] = {"date_style": style}
        lines[style] = csv_data_line(event, user_with_perms, columns)

    assert lines["iso"] != lines["date_only"], "both styles produced %r" % (
        lines["iso"],
    )
    assert "T" in lines["iso"]  # ISO 8601 separates date and time with a T
    assert "T" not in lines["date_only"]
    assert ":" not in lines["date_only"]  # no time of day at all


@pytest.mark.django_db
def test_a_number_style_reaches_the_export(registered, event, user_with_perms, orders):
    """``currency`` and ``localized`` become strings, ``raw`` stays a number.

    The three cases in one row of one report, because the interesting property
    is that they are *different from each other* -- a renderer that formatted
    everything, or nothing, would satisfy any one of them alone.
    """
    row = first_data_row(
        event,
        user_with_perms,
        [
            {"field": "order.total"},
            {"field": "order.total", "format": {"number_style": "currency"}},
            {"field": "order.total", "format": {"number_style": "localized"}},
            {"field": "order.total", "format": {"number_style": "raw"}},
        ],
    )
    plain, currency, localized, raw = row

    assert plain == Decimal("23.00")
    assert isinstance(plain, Decimal), "no format must not touch the value"
    assert isinstance(currency, str) and "23" in currency
    assert currency != str(plain), "currency formatting did not happen"
    assert isinstance(localized, str) and "23" in localized
    assert isinstance(raw, Decimal), "raw is the one style that keeps the type"


@pytest.mark.django_db
def test_a_boolean_style_reaches_the_export(registered, event, user_with_perms, orders):
    """``order.testmode`` is ``False`` for order AAAAA -- four ways of saying so."""
    row = first_data_row(
        event,
        user_with_perms,
        [
            {"field": "order.testmode"},
            {"field": "order.testmode", "format": {"boolean_style": "one_zero"}},
            {"field": "order.testmode", "format": {"boolean_style": "true_false"}},
            {"field": "order.testmode", "format": {"boolean_style": "yes_no"}},
        ],
    )
    plain, one_zero, true_false, yes_no = row

    assert plain is False, "no format must not touch the value"
    assert one_zero == "0"
    assert true_false == "false"
    # Translated, so this asserts the shape rather than the English word, the
    # same way the preview test in tests/test_editor_api.py does.
    assert yes_no not in ("0", "false", "False")
    assert isinstance(yes_no, str) and yes_no


@pytest.mark.django_db
def test_a_boolean_style_is_visible_in_the_csv_bytes(
    registered, event, user_with_perms, orders
):
    """Not only in the row list: the file itself has to contain it.

    ``_render_csv`` writes ``False`` as ``False`` and ``"0"`` as ``"0"``; this
    is the assertion that nothing between the renderer and the bytes undoes the
    formatting.
    """
    line = csv_data_line(
        event,
        user_with_perms,
        [{"field": "order.testmode", "format": {"boolean_style": "one_zero"}}],
    )
    assert line.strip('"') == "0"
    assert "False" not in line


@pytest.mark.django_db
def test_a_report_without_a_column_format_is_exported_exactly_as_before(
    registered, event, user_with_perms, orders
):
    """The control group: every report saved before T-001 keeps its old file.

    Two assertions, because "unchanged" has two halves. The row still carries
    the native types the compiler produced -- and the exporter does not even
    build a formatting pass for such a report, so the old path is literally the
    old path rather than a formatter that happens to be a no-op.
    """
    columns = [
        {"field": "order.code"},
        {"field": "order.datetime"},
        {"field": "order.total"},
        {"field": "order.testmode"},
    ]
    row = first_data_row(event, user_with_perms, columns)
    assert row[0] == "AAAAA"
    assert isinstance(row[1], datetime.datetime)
    assert isinstance(row[2], Decimal)
    assert row[3] is False

    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        _report, compiled = ex._prepare(event, "fmt", {})
        assert ex._cell_formats(compiled) is None

    # A ``separator`` is not a style: the compiler applies it, so it must not
    # switch the formatting pass on either.
    with scopes_disabled():
        ReportDefinition.objects.filter(event=event, identifier="sep").delete()
        make_format_report(
            event,
            [{"field": "order.code", "format": {"separator": " / "}}],
            identifier="sep",
        )
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        _report, compiled = ex._prepare(event, "sep", {})
        assert ex._cell_formats(compiled) is None


@pytest.mark.django_db
def test_a_hidden_column_does_not_shift_the_formats(
    registered, event, user_with_perms, orders
):
    """Formats are paired with the *visible* columns, like in the preview.

    ``CompiledReport.columns`` has hidden columns dropped already. Pairing by
    the raw definition index would hand column 0's (absent) format to the first
    output column and silently drop the style the user set -- the T-001 bug
    again, one layer down and much harder to see.
    """
    row = first_data_row(
        event,
        user_with_perms,
        [
            {"field": "order.code", "hidden": True},
            {"field": "order.testmode", "format": {"boolean_style": "one_zero"}},
        ],
    )
    assert row == ["0"]


@pytest.mark.django_db
def test_number_style_raw_keeps_a_native_number_in_the_xlsx(
    registered, event, user_with_perms, orders
):
    """The reason formatting is not in the compiler, asserted in a real file.

    A spreadsheet can add up column 1 and cannot add up column 2. If the
    renderer stringified everything, both would be text and ``NumberStyle.RAW``
    would be a lie (its own docstring promises the opposite).
    """
    import io
    from openpyxl import load_workbook

    with scopes_disabled():
        make_format_report(
            event,
            [
                {"field": "order.total", "format": {"number_style": "raw"}},
                {"field": "order.total", "format": {"number_style": "currency"}},
                {"field": "order.datetime", "format": {"date_style": "date_only"}},
            ],
        )
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        _fn, _mime, content = render_xlsx(
            ex, {"_format": "xlsx", contracts.EXPORT_FORM_REPORT_KEY: "fmt"}
        )
    sheet = load_workbook(io.BytesIO(content)).active
    raw, currency, date_only = (sheet.cell(row=2, column=i).value for i in (1, 2, 3))

    assert isinstance(raw, (int, float, Decimal)) and not isinstance(raw, bool)
    assert float(raw) == 23.0
    assert isinstance(currency, str) and "23" in currency
    assert isinstance(date_only, str) and ":" not in date_only


# ---------------------------------------------------------------------------
# 2c. Timezones and XLSX
# ---------------------------------------------------------------------------
#
# Found while fixing T-001 and fixed in the same file: openpyxl refuses a
# timezone-aware datetime outright, so *every* XLSX export of a report with a
# date column used to raise TypeError -- not an ExportError, so five Celery
# retries and the word "Internal Error". Only the XLSX path is affected; CSV
# writes an aware datetime correctly and is deliberately left alone.


def xlsx_sheet(event, user_with_perms, columns, identifier="fmt"):
    """Render a report to XLSX and hand back the loaded worksheet."""
    import io
    from openpyxl import load_workbook

    with scopes_disabled():
        ReportDefinition.objects.filter(event=event, identifier=identifier).delete()
        make_format_report(event, columns, identifier=identifier)
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        _fn, _mime, content = render_xlsx(
            ex, {"_format": "xlsx", contracts.EXPORT_FORM_REPORT_KEY: identifier}
        )
    return load_workbook(io.BytesIO(content)).active


@pytest.mark.django_db
def test_a_datetime_column_without_a_style_survives_the_xlsx_path(
    registered, event, user_with_perms, orders
):
    """``Excel does not support timezones in datetimes``, says openpyxl.

    ``_render_xlsx`` appends our cell values unchanged (pretix/base/exporter.py
    :305-311) and ``SafeCell`` passes a ``datetime`` straight to openpyxl, which
    rejects an aware one outright. Our rows are aware -- ``Order.datetime`` is a
    ``DateTimeField`` and ``USE_TZ`` is on.

    Two assertions, because "it does not crash" is only half of it: the cell has
    to be a real ``datetime`` (a spreadsheet can sort and subtract it), and it
    has to show the event's local wall clock rather than UTC. The event here is
    deliberately not in UTC, so the two are three or four digits apart.
    """
    with scopes_disabled():
        event.settings.timezone = "Europe/Berlin"
        order = Order.objects.get(event=event, code="AAAAA")

    sheet = xlsx_sheet(event, user_with_perms, [{"field": "order.datetime"}])
    cell = sheet.cell(row=2, column=1)

    assert isinstance(cell.value, datetime.datetime)
    assert cell.value.tzinfo is None
    # Compared to the second: XLSX stores a datetime as a serial number of days,
    # so the microseconds of the stored order do not survive the round trip.
    expected = order.datetime.astimezone(ZoneInfo("Europe/Berlin"))
    assert cell.value.replace(microsecond=0) == expected.replace(
        tzinfo=None, microsecond=0
    )
    assert cell.value.replace(microsecond=0) != order.datetime.replace(
        tzinfo=None, microsecond=0
    ), "written in UTC instead of the event's timezone"


@pytest.mark.django_db
def test_without_the_naive_conversion_the_xlsx_path_fails_at_openpyxl(
    registered, event, user_with_perms, orders, monkeypatch
):
    """The counter-check: with the fix switched off, the old error comes back.

    Not "some export fails" -- *this* error, from openpyxl, about timezones. A
    test that only asserted "does not raise" would stay green if the datetime
    column silently disappeared, and it is the fix, not the report, that has to
    be doing the work.
    """
    import io

    monkeypatch.setattr(exporters, "as_spreadsheet_value", lambda value, event: value)
    with scopes_disabled():
        make_format_report(event, [{"field": "order.datetime"}])
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        with pytest.raises(TypeError) as excinfo:
            ex.render(
                {"_format": "xlsx", contracts.EXPORT_FORM_REPORT_KEY: "fmt"},
                output_file=io.BytesIO(),
            )
    assert "timezone" in str(excinfo.value).lower()


@pytest.mark.django_db
def test_a_chosen_date_style_is_unaffected_by_the_xlsx_conversion(
    registered, event, user_with_perms, orders
):
    """Control group: the T-001 styles keep working, unchanged, in XLSX.

    A styled cell is a string by the time the spreadsheet conversion sees it,
    so it must pass through untouched -- and it must be the *same* string the
    CSV export produces, which is the whole point of the shared renderer.
    """
    with scopes_disabled():
        event.settings.timezone = "Europe/Berlin"

    columns = [
        {"field": "order.datetime", "format": {"date_style": "iso"}},
        {"field": "order.datetime", "format": {"date_style": "date_only"}},
    ]
    sheet = xlsx_sheet(event, user_with_perms, columns)
    iso, date_only = (sheet.cell(row=2, column=i).value for i in (1, 2))

    assert isinstance(iso, str) and "T" in iso
    assert iso.endswith("+02:00") or iso.endswith("+01:00"), iso
    assert isinstance(date_only, str) and ":" not in date_only

    csv_row = first_data_row(event, user_with_perms, columns)
    assert csv_row == [iso, date_only]


@pytest.mark.django_db
def test_the_csv_path_still_writes_the_timezone(
    registered, event, user_with_perms, orders
):
    """The other half of the control group: CSV bytes do not change at all.

    The fix is XLSX-only on purpose. An aware datetime is written correctly by
    ``csv.writer``, and making it naive here would rewrite every existing
    report's file for no gain.
    """
    with scopes_disabled():
        event.settings.timezone = "Europe/Berlin"
    line = csv_data_line(event, user_with_perms, [{"field": "order.datetime"}])
    assert "+00:00" in line, line  # unchanged: aware, and still UTC


def test_as_spreadsheet_value_touches_nothing_else():
    """Only aware temporal values change, and only in their tzinfo."""
    naive = datetime.datetime(2026, 3, 1, 9, 30)
    assert exporters.as_spreadsheet_value(naive) is naive
    assert exporters.as_spreadsheet_value(datetime.date(2026, 3, 1)) == datetime.date(
        2026, 3, 1
    )
    assert exporters.as_spreadsheet_value(Decimal("23.50")) == Decimal("23.50")
    assert exporters.as_spreadsheet_value("2026-03-01T09:30:00+00:00") == (
        "2026-03-01T09:30:00+00:00"
    )
    assert exporters.as_spreadsheet_value(None) is None
    assert exporters.as_spreadsheet_value(True) is True

    aware_time = datetime.time(9, 30, tzinfo=datetime.timezone.utc)
    assert exporters.as_spreadsheet_value(aware_time) == datetime.time(9, 30)
    assert exporters.as_spreadsheet_value(datetime.time(9, 30)) == datetime.time(9, 30)


def test_as_spreadsheet_value_without_an_event_does_not_crash():
    """A missing or broken timezone must cost a wrong hour, not the export."""
    aware = datetime.datetime(2026, 3, 1, 9, 30, tzinfo=datetime.timezone.utc)
    assert exporters.as_spreadsheet_value(aware, None) == datetime.datetime(
        2026, 3, 1, 9, 30
    )


# --- the renderer itself, without a database in the way ---------------------


def test_format_export_cell_hands_unstyled_values_through_untouched():
    """The four "do not touch this" cases of the export policy.

    Each one is somebody's promise: old reports keep their files, ``RAW`` keeps
    a spreadsheet number, ``None`` keeps an empty XLSX cell instead of a string,
    and a value the compiler already rendered is not rendered a second time.
    """
    when = datetime.datetime(2026, 3, 1, 9, 30, tzinfo=datetime.timezone.utc)

    # No format at all.
    assert exporters.format_export_cell(when, None) is when
    assert exporters.format_export_cell(Decimal("23.50"), None) == Decimal("23.50")
    assert exporters.format_export_cell(True, None) is True

    # A format that says nothing about this type.
    only_separator = ColumnFormat(separator=" / ")
    assert exporters.format_export_cell(when, only_separator) is when
    number_only = ColumnFormat(number_style=NumberStyle.CURRENCY)
    assert exporters.format_export_cell(when, number_only) is when
    assert exporters.format_export_cell(True, number_only) is True

    # RAW: the one style whose meaning is "leave me alone".
    raw = ColumnFormat(number_style=NumberStyle.RAW)
    assert exporters.format_export_cell(Decimal("23.50"), raw) == Decimal("23.50")
    assert isinstance(exporters.format_export_cell(Decimal("23.50"), raw), Decimal)

    # None and str, whatever the format says.
    iso = ColumnFormat(date_style=DateStyle.ISO)
    assert exporters.format_export_cell(None, iso) is None
    assert exporters.format_export_cell("2026", iso) == "2026"


def test_format_export_cell_does_apply_the_three_styles():
    """The other half of the policy, so the test above cannot pass vacuously."""
    when = datetime.datetime(2026, 3, 1, 9, 30, tzinfo=datetime.timezone.utc)
    assert exporters.format_export_cell(
        when, ColumnFormat(date_style=DateStyle.ISO)
    ).startswith("2026-03-01T")
    assert (
        exporters.format_export_cell(
            Decimal("23.50"), ColumnFormat(number_style=NumberStyle.LOCALIZED)
        )
        == "23.50"
    )
    assert (
        exporters.format_export_cell(
            True, ColumnFormat(boolean_style=BooleanStyle.ONE_ZERO)
        )
        == "1"
    )


def test_a_style_that_does_not_fit_the_value_is_not_a_celery_crash():
    """A definition is untrusted input, and nothing revalidates it per run.

    ``date_only`` on a time-of-day value asks Django for the day number of a
    ``datetime.time``. The export must survive that with an unformatted cell:
    an ``AttributeError`` here would leave this module as something other than
    an ``ExportError``, which costs five retries and produces the word
    "Internal Error" (services/export.py:392-397).
    """
    value = datetime.time(9, 30)
    fmt = ColumnFormat(date_style=DateStyle.DATE_ONLY)
    with pytest.raises(Exception):
        exporters.format_cell_value(value, fmt)  # the strict renderer does raise
    assert exporters.format_export_cell(value, fmt) is value  # the export does not


@pytest.mark.django_db
def test_the_preview_and_the_export_share_one_renderer(event):
    """The preview must not be prettier than the export -- as an equality.

    Two ways to satisfy this, and both are accepted: ``views/api.py`` imports
    :func:`~pretix_custom_reports.exporters.format_cell_value` (the intended end
    state, frontend-dev's follow-up round), or it still has its own function --
    in which case the two are compared cell by cell over the whole matrix of
    value types and styles. What is *not* accepted is two renderers that
    disagree.
    """
    from pretix_custom_reports.views import api

    preview_renderer = getattr(api, "format_cell", None)
    if preview_renderer is None or preview_renderer is exporters.format_cell_value:
        return  # deduplicated; there is only one renderer left to compare

    values = [
        None,
        "text",
        True,
        False,
        Decimal("23.50"),
        17,
        4.5,
        datetime.datetime(2026, 3, 1, 9, 30, tzinfo=datetime.timezone.utc),
        datetime.date(2026, 3, 1),
        datetime.time(9, 30),
    ]
    column_formats = [None, ColumnFormat()]
    column_formats += [ColumnFormat(date_style=style) for style in DateStyle]
    column_formats += [ColumnFormat(number_style=style) for style in NumberStyle]
    column_formats += [ColumnFormat(boolean_style=style) for style in BooleanStyle]
    datatypes = [None, DataType.MONEY, DataType.DATETIME, DataType.BOOLEAN]

    def outcome(fn, *args):
        try:
            return ("ok", fn(*args))
        except Exception as e:  # both may legitimately refuse the same input
            return ("raised", type(e).__name__)

    compared = 0
    for value in values:
        for fmt in column_formats:
            for datatype in datatypes:
                mine = outcome(exporters.format_cell_value, value, fmt, datatype, event)
                theirs = outcome(preview_renderer, value, fmt, datatype, event)
                assert mine == theirs, (value, fmt, datatype, mine, theirs)
                compared += 1
    assert compared == len(values) * len(column_formats) * len(datatypes)


# ---------------------------------------------------------------------------
# 3. Multi-event
# ---------------------------------------------------------------------------


@pytest.fixture
def two_events(event):
    """A second event in the same organizer, with its own order."""
    with scopes_disabled():
        second = Event.objects.create(
            organizer=event.organizer,
            name="Second Event",
            slug="second",
            date_from=now() + datetime.timedelta(days=60),
            plugins="pretix_custom_reports",
            live=True,
        )
        channel = event.organizer.sales_channels.get(identifier="web")
        item = Item.objects.create(
            event=second, name="Ticket", internal_name="ticket", default_price=5
        )
        order = Order.objects.create(
            event=second,
            code="ZZZZZ",
            status=Order.STATUS_PAID,
            email="z@example.org",
            sales_channel=channel,
            datetime=now() - datetime.timedelta(days=1),
            expires=now() + datetime.timedelta(days=10),
            total=Decimal("5.00"),
        )
        OrderPosition.objects.create(
            order=order, item=item, price=Decimal("5.00"), positionid=1
        )
        return second


@pytest.mark.django_db
def test_multievent_export_covers_every_event_and_names_it(
    registered, event, two_events, user_with_perms, orders
):
    with scopes_disabled():
        make_report(event=event, identifier="codes")
        make_report(event=two_events, identifier="codes")
    with scope(organizer=event.organizer):
        ex = exporter_for_organizer(event.organizer, user_with_perms)
        rows = rows_of(ex, csv_form_data())
    assert rows[0][:2] == ["Event slug", "Event name"]
    body = {(row[0], row[2]) for row in rows[1:]}
    assert ("dummy", "AAAAA") in body
    assert ("second", "ZZZZZ") in body


@pytest.mark.django_db
def test_an_event_whose_report_is_missing_is_skipped_not_fatal(
    registered, event, two_events, user_with_perms, orders
):
    """The report exists in one event only. The export must still deliver."""
    with scopes_disabled():
        make_report(event=event, identifier="codes")
    with scope(organizer=event.organizer):
        ex = exporter_for_organizer(event.organizer, user_with_perms)
        rows = rows_of(ex, csv_form_data())
    slugs = {row[0] for row in rows[1:]}
    assert slugs == {"dummy"}


@pytest.mark.django_db
def test_an_unresolvable_field_in_one_event_is_skipped_not_fatal(
    registered, event, two_events, user_with_perms, orders
):
    """A question exists in event A and not in event B -- SPEC.md F9's daily case.

    The compiler raises ``FieldResolutionError`` for event B. Without the
    translation into a per-event skip, the whole organizer export would die and
    the user would never see event A's rows.
    """
    with scopes_disabled():
        question = Question.objects.create(
            event=event,
            question="T-shirt size",
            identifier="tshirt-size",
            type=Question.TYPE_TEXT,
            position=0,
        )
        position = OrderPosition.objects.filter(order__event=event).first()
        QuestionAnswer.objects.create(
            orderposition=position, question=question, answer="L"
        )
        definition = {
            "schema_version": contracts.SCHEMA_VERSION,
            "base": "orderposition",
            "columns": [
                {"field": "order.code"},
                {"field": "answer.tshirt-size"},
            ],
        }
        for target in (event, two_events):
            make_report(event=target, identifier="withquestion", definition=definition)
    with scope(organizer=event.organizer):
        ex = exporter_for_organizer(event.organizer, user_with_perms)
        rows = rows_of(ex, csv_form_data(identifier="withquestion"))
        # ... and with the strict policy, the message has to name the field, not
        # just the event. Otherwise "skipped" and "report missing" would be
        # indistinguishable to whoever gets the mail.
        with pytest.raises(ExportError) as excinfo:
            ex.render(
                csv_form_data(
                    identifier="withquestion",
                    **{
                        exporters.FORM_KEY_ON_UNAVAILABLE: exporters.ON_UNAVAILABLE_FAIL
                    },
                )
            )
    assert {row[0] for row in rows[1:]} == {"dummy"}
    assert ["L"] == [row[3] for row in rows[1:] if row[3]]
    assert "answer.tshirt-size" in str(excinfo.value)
    assert "second" in str(excinfo.value)


@pytest.mark.django_db
def test_fail_policy_turns_the_same_situation_into_an_exporterror(
    registered, event, two_events, user_with_perms, orders
):
    """Skipping is the default, not the only option.

    Silently dropping a whole event from a report someone bills against is its
    own kind of wrong answer, so the choice is explicit in the form.
    """
    with scopes_disabled():
        make_report(event=event, identifier="codes")
    with scope(organizer=event.organizer):
        ex = exporter_for_organizer(event.organizer, user_with_perms)
        with pytest.raises(ExportError) as excinfo:
            ex.render(
                csv_form_data(
                    **{exporters.FORM_KEY_ON_UNAVAILABLE: exporters.ON_UNAVAILABLE_FAIL}
                )
            )
    assert "second" in str(excinfo.value)


@pytest.mark.django_db
def test_if_no_event_can_supply_the_report_the_export_fails_loudly(
    registered, event, two_events, user_with_perms, orders
):
    """Not an empty file.

    An empty result becomes ``ExportEmptyError``, which pretix treats as a
    *soft* failure: mail sent, error counter untouched (services/export.py:
    371-374, 388-389). A schedule pointing at a deleted report would then keep
    mailing "no data" forever without ever naming the cause.
    """
    with scope(organizer=event.organizer):
        ex = exporter_for_organizer(event.organizer, user_with_perms)
        with pytest.raises(ExportError) as excinfo:
            ex.render(csv_form_data(identifier="ghost"))
    assert "ghost" in str(excinfo.value)


@pytest.mark.django_db
def test_two_reports_sharing_an_identifier_but_not_a_shape_do_not_produce_ragged_rows(
    registered, event, two_events, user_with_perms, orders
):
    """Identifiers are unique per event, so this really can happen."""
    with scopes_disabled():
        make_report(event=event, identifier="codes")
        make_report(
            event=two_events,
            identifier="codes",
            definition={
                "schema_version": contracts.SCHEMA_VERSION,
                "base": "order",
                "columns": [{"field": "order.code"}, {"field": "order.email"}],
            },
        )
    with scope(organizer=event.organizer):
        ex = exporter_for_organizer(event.organizer, user_with_perms)
        rows = rows_of(ex, csv_form_data())
    widths = {len(row) for row in rows}
    assert len(widths) == 1
    assert {row[0] for row in rows[1:]} == {"dummy"}


@pytest.mark.django_db
def test_a_single_event_export_never_skips(registered, event, user_with_perms, orders):
    """At event level there is nothing left to export, so skipping is a lie."""
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        with pytest.raises(ExportError) as excinfo:
            ex.render(
                csv_form_data(
                    identifier="ghost",
                    **{
                        exporters.FORM_KEY_ON_UNAVAILABLE: exporters.ON_UNAVAILABLE_SKIP
                    },
                )
            )
    assert "ghost" in str(excinfo.value)
    assert "dummy" in str(excinfo.value)


# ---------------------------------------------------------------------------
# 3b. The plugin gate (security review S-002)
# ---------------------------------------------------------------------------
#
# On event level pretix gates us itself: ``register_data_exporters`` is an
# ``EventPluginSignal`` and does not fire for an event without the plugin, so
# ``init_event_exporters`` never builds us there (asserted in
# tests/test_security.py, which owns that control group).
#
# On organizer level nothing gates us. ``register_multievent_data_exporters`` is
# an ``OrganizerPluginSignal(allow_legacy_plugins=True)``, so an event-level
# plugin counts as active for *every* organizer (pretix/base/signals.py:100-113),
# and ``self.events`` is filtered by permission only
# (``init_organizer_exporters``, services/export.py:266-287). Until S-002 that
# meant an event whose administrator had switched the plugin off still delivered
# its orders into the organizer export file -- and through
# ``ScheduledOrganizerExport`` into a recurring mail nobody was reviewing.
#
# The gate is ``self.plugin_module in event.get_plugins()``, per event, applied
# in ``report_choices`` and at the very top of ``_prepare``.


def disable_plugin(event):
    """Switch the plugin off for *event*, the way the plugin page does."""
    with scopes_disabled():
        event.plugins = ""
        event.save(update_fields=["plugins"])
    return event


@pytest.mark.django_db
def test_an_event_with_the_plugin_switched_off_contributes_no_rows(
    registered, event, two_events, user_with_perms, orders
):
    """The finding itself: order data of a switched-off event in the file.

    Both events hold the report and both hold orders, so the only thing that can
    keep ``ZZZZZ`` out of the result is the plugin gate. The rest of the export
    has to keep working -- one deactivated event must not cost an organizer the
    other four.
    """
    with scopes_disabled():
        make_report(event=event, identifier="codes")
        make_report(event=two_events, identifier="codes")
    disable_plugin(two_events)

    with scope(organizer=event.organizer):
        ex = exporter_for_organizer(event.organizer, user_with_perms)
        rows = rows_of(ex, csv_form_data())

    assert {row[0] for row in rows[1:]} == {"dummy"}
    assert "ZZZZZ" not in {row[2] for row in rows[1:]}
    assert "AAAAA" in {row[2] for row in rows[1:]}


@pytest.mark.django_db
def test_the_switched_off_event_is_skipped_through_the_documented_mechanism(
    registered, event, two_events, user_with_perms, orders
):
    """Not a silent drop: the same ``_EventProblem`` path as a deleted report.

    Which matters twice. Under the strict policy the export fails and the
    message has to say *why* an event is missing -- "the plugin is not enabled"
    and "the report does not exist there" call for entirely different repairs.
    And when nothing at all can be exported, the reason has to survive into
    ``ExportError``, because that string is what the schedule owner receives by
    mail.
    """
    with scopes_disabled():
        make_report(event=event, identifier="codes")
        make_report(event=two_events, identifier="codes")
    disable_plugin(two_events)

    with scope(organizer=event.organizer):
        ex = exporter_for_organizer(event.organizer, user_with_perms)
        with pytest.raises(ExportError) as strict:
            ex.render(
                csv_form_data(
                    **{exporters.FORM_KEY_ON_UNAVAILABLE: exporters.ON_UNAVAILABLE_FAIL}
                )
            )
    message = str(strict.value)
    assert "second" in message
    assert "plugin is not enabled" in message

    # And with every event switched off there is no file, not an empty one.
    disable_plugin(event)
    with scope(organizer=event.organizer):
        ex = exporter_for_organizer(event.organizer, user_with_perms)
        with pytest.raises(ExportError) as everything:
            ex.render(csv_form_data())
    assert "plugin is not enabled" in str(everything.value)
    assert "dummy" in str(everything.value)


@pytest.mark.django_db
def test_the_plugin_gate_comes_before_the_report_lookup(
    registered, event, two_events, user_with_perms, orders
):
    """A switched-off event is refused for a reason of its own.

    Order of the two checks in ``_prepare`` is observable: here the report does
    exist in the deactivated event, so if the lookup ran first the message would
    talk about columns or produce rows. The plugin is the more fundamental
    answer and has to be the one reported.
    """
    with scopes_disabled():
        make_report(event=two_events, identifier="only-there")
    disable_plugin(two_events)

    with scope(organizer=event.organizer):
        ex = exporter_for_organizer(event.organizer, user_with_perms)
        with pytest.raises(ExportError) as excinfo:
            ex.render(csv_form_data(identifier="only-there"))
    message = str(excinfo.value)
    assert "plugin is not enabled" in message
    assert "second" in message


@pytest.mark.django_db
def test_report_choices_hide_reports_of_events_with_the_plugin_switched_off(
    registered, event, two_events, user_with_perms
):
    """Offering a choice that ``_prepare`` will refuse is its own bug."""
    with scopes_disabled():
        make_report(event=event, identifier="codes")
        make_report(event=two_events, identifier="only-there", name="Only there")

    with scope(organizer=event.organizer):
        before = dict(
            exporter_for_organizer(event.organizer, user_with_perms).report_choices()
        )
    assert set(before) == {"codes", "only-there"}

    disable_plugin(two_events)
    with scope(organizer=event.organizer):
        after = dict(
            exporter_for_organizer(event.organizer, user_with_perms).report_choices()
        )
    assert set(after) == {"codes"}


@pytest.mark.django_db
def test_the_plugin_gate_is_not_a_substring_match(
    registered, event, two_events, user_with_perms, orders
):
    """``plugins__contains`` would have accepted a longer, foreign plugin name.

    ``Event.plugins`` is one comma-separated string, so a ``contains`` filter
    matches any plugin whose name merely starts with ours. pretix itself splits
    the list and compares whole entries (``get_plugins()``,
    base/models/event.py:794-800), and so do we.
    """
    with scopes_disabled():
        make_report(event=event, identifier="codes")
        make_report(event=two_events, identifier="codes")
        two_events.plugins = "pretix_custom_reports_extra"
        two_events.save(update_fields=["plugins"])

    with scope(organizer=event.organizer):
        ex = exporter_for_organizer(event.organizer, user_with_perms)
        rows = rows_of(ex, csv_form_data())
    assert {row[0] for row in rows[1:]} == {"dummy"}


@pytest.mark.django_db
def test_a_neighbouring_plugin_in_the_same_list_does_not_disable_us(
    registered, event, two_events, user_with_perms, orders
):
    """The counter-test to the one above -- the gate must not be too narrow.

    An event normally carries several plugins, and ours can sit anywhere in that
    comma-separated list.
    """
    with scopes_disabled():
        make_report(event=event, identifier="codes")
        make_report(event=two_events, identifier="codes")
        two_events.plugins = "pretix.plugins.banktransfer,pretix_custom_reports"
        two_events.save(update_fields=["plugins"])

    with scope(organizer=event.organizer):
        ex = exporter_for_organizer(event.organizer, user_with_perms)
        rows = rows_of(ex, csv_form_data())
    assert {row[0] for row in rows[1:]} == {"dummy", "second"}


@pytest.mark.django_db
def test_a_scheduled_organizer_export_stops_sending_a_deactivated_event(
    registered, event, two_events, schedule_user, orders
):
    """The reason S-002 is worth more than a UI nit.

    A ``ScheduledOrganizerExport`` keeps mailing on its own; a schedule created
    while both events ran the plugin must not keep attaching the orders of an
    event that has since been switched off. This is the same run the finding was
    measured on, only through the scheduler instead of ``render``.
    """
    djmail.outbox = []
    with scopes_disabled():
        make_report(event=event, identifier="codes")
        make_report(event=two_events, identifier="codes")
        schedule = ScheduledOrganizerExport(
            organizer=event.organizer, owner=schedule_user
        )
        schedule.export_identifier = exporters.CustomReportExporter.identifier
        schedule.export_form_data = {
            contracts.EXPORT_FORM_REPORT_KEY: "codes",
            "_format": "default",
            "all_events": True,
        }
        schedule.locale = "en"
        schedule.mail_subject = "All events"
        schedule.mail_template = "Here you go."
        schedule.schedule_rrule = (
            "DTSTART:20260118T000000\nRRULE:FREQ=DAILY;INTERVAL=1;WKST=MO"
        )
        schedule.schedule_rrule_time = datetime.time(2, 30, 0)
        schedule.schedule_next_run = now() - datetime.timedelta(minutes=5)
        schedule.save()
    disable_plugin(two_events)

    run_scheduled_exports(None)

    schedule.refresh_from_db()
    assert schedule.error_counter == 0, schedule.error_last_message
    payload = djmail.outbox[0].attachments[0][1]
    if isinstance(payload, bytes):
        payload = payload.decode("utf-8")
    assert "AAAAA" in payload
    assert "ZZZZZ" not in payload


# ---------------------------------------------------------------------------
# 4. Scheduled exports -- the deleted report case
# ---------------------------------------------------------------------------


@pytest.fixture
def schedule_user(organizer):
    user = User.objects.create_user("owner@example.org", "dummy")
    team = Team.objects.create(
        organizer=organizer,
        name="Owner team",
        all_events=True,
        all_event_permissions=True,
        all_organizer_permissions=True,
    )
    team.members.add(user)
    return user


def make_schedule(event, owner, identifier="codes", **form_data):
    data = {contracts.EXPORT_FORM_REPORT_KEY: identifier, "_format": "default"}
    data.update(form_data)
    schedule = ScheduledEventExport(event=event, owner=owner)
    schedule.export_identifier = exporters.CustomReportExporter.identifier
    schedule.export_form_data = data
    schedule.locale = "en"
    schedule.mail_subject = "Your report"
    schedule.mail_template = "Here you go."
    schedule.schedule_rrule = (
        "DTSTART:20260118T000000\nRRULE:FREQ=DAILY;INTERVAL=1;WKST=MO"
    )
    schedule.schedule_rrule_time = datetime.time(2, 30, 0)
    schedule.schedule_next_run = now() - datetime.timedelta(minutes=5)
    schedule.save()
    return schedule


@pytest.mark.django_db
def test_a_scheduled_export_can_be_created_and_runs(
    registered, event, schedule_user, orders
):
    """The end-to-end proof that we hang off pretix' scheduler, not our own.

    ``run_scheduled_exports`` is the ``periodic_task`` receiver; with
    ``CELERY_TASK_ALWAYS_EAGER`` the dispatched task runs inline, so this
    exercises ``EventTask`` (which opens the django-scopes scope),
    ``init_event_exporter`` (which re-checks the owner's permission) and
    ``_run_scheduled_export`` (which mails the file).
    """
    djmail.outbox = []
    with scopes_disabled():
        make_report(event=event, identifier="codes")
        schedule = make_schedule(event, schedule_user)

    run_scheduled_exports(None)

    schedule.refresh_from_db()
    assert schedule.error_counter == 0
    assert schedule.schedule_next_run > now()
    assert len(djmail.outbox) == 1
    assert djmail.outbox[0].subject == "Your report"
    attachment = djmail.outbox[0].attachments[0]
    assert attachment[0].endswith(".csv")
    payload = attachment[1]
    if isinstance(payload, bytes):
        payload = payload.decode("utf-8")
    assert "AAAAA" in payload


@pytest.mark.django_db
def test_a_scheduled_export_whose_report_was_deleted_explains_itself(
    registered, event, schedule_user, orders
):
    """The failure this module exists for.

    Without the ``ObjectDoesNotExist`` branch in ``_prepare``, the
    ``DoesNotExist`` would reach ``services/export.py:392-397``: five Celery
    retries at 120 s, then a mail saying "Internal Error", then -- after five
    such runs -- the schedule disappears from the periodic query without another
    word. What we want instead is one mail, immediately, naming the report and
    the event.
    """
    djmail.outbox = []
    with scopes_disabled():
        report = make_report(event=event, identifier="codes")
        schedule = make_schedule(event, schedule_user)
        report.delete()

    run_scheduled_exports(None)

    schedule.refresh_from_db()
    assert schedule.error_counter == 1
    assert "codes" in schedule.error_last_message
    assert event.slug in schedule.error_last_message
    assert "Internal Error" not in schedule.error_last_message
    assert len(djmail.outbox) == 1
    assert djmail.outbox[0].subject == "Export failed"
    assert "codes" in djmail.outbox[0].body


@pytest.mark.django_db
def test_the_deleted_report_error_reaches_the_event_log(
    registered, event, schedule_user, orders
):
    """The failure is recorded where an administrator will find it later."""
    with scopes_disabled():
        report = make_report(event=event, identifier="codes")
        make_schedule(event, schedule_user)
        report.delete()

    run_scheduled_exports(None)

    with scopes_disabled():
        entry = (
            event.logentry_set.filter(action_type="pretix.event.export.schedule.failed")
            .order_by("-pk")
            .first()
        )
    assert entry is not None
    assert "codes" in entry.parsed_data["reason"]
    assert entry.parsed_data["soft"] is False


@pytest.mark.django_db
def test_a_scheduled_organizer_export_runs_over_several_events(
    registered, event, two_events, schedule_user, orders
):
    """``ScheduledOrganizerExport`` plus the ``all_events`` marker of the UI.

    ``all_events``/``events`` are injected by the organizer export view, not by
    the exporter (api-notes section 5.3); the task filters on them before we are
    constructed, which is exactly why we must read only ``self.events``.
    """
    djmail.outbox = []
    with scopes_disabled():
        make_report(event=event, identifier="codes")
        make_report(event=two_events, identifier="codes")
        schedule = ScheduledOrganizerExport(
            organizer=event.organizer, owner=schedule_user
        )
        schedule.export_identifier = exporters.CustomReportExporter.identifier
        schedule.export_form_data = {
            contracts.EXPORT_FORM_REPORT_KEY: "codes",
            "_format": "default",
            "all_events": True,
        }
        schedule.locale = "en"
        schedule.mail_subject = "All events"
        schedule.mail_template = "Here you go."
        schedule.schedule_rrule = (
            "DTSTART:20260118T000000\nRRULE:FREQ=DAILY;INTERVAL=1;WKST=MO"
        )
        schedule.schedule_rrule_time = datetime.time(2, 30, 0)
        schedule.schedule_next_run = now() - datetime.timedelta(minutes=5)
        schedule.save()

    run_scheduled_exports(None)

    schedule.refresh_from_db()
    assert schedule.error_counter == 0, schedule.error_last_message
    payload = djmail.outbox[0].attachments[0][1]
    if isinstance(payload, bytes):
        payload = payload.decode("utf-8")
    assert "AAAAA" in payload
    assert "ZZZZZ" in payload


@pytest.mark.django_db
def test_a_scheduled_export_of_an_inactive_owner_is_refused_by_pretix(
    registered, event, schedule_user, orders
):
    """Background runs use the owner's permissions, not nobody's.

    pretix checks ``owner.is_active`` and re-runs ``init_event_exporter`` with
    the owner as the acting account (services/export.py:471-477). This test is
    here so that a later "let us just run it as a superuser" refactor of ours
    would be caught.
    """
    with scopes_disabled():
        make_report(event=event, identifier="codes")
        schedule = make_schedule(event, schedule_user)
        schedule_user.is_active = False
        schedule_user.save()

    run_scheduled_exports(None)

    schedule.refresh_from_db()
    assert schedule.error_counter == 1


# ---------------------------------------------------------------------------
# 5. Relative date filters are evaluated per run
# ---------------------------------------------------------------------------


RELATIVE_DEFINITION = {
    "schema_version": contracts.SCHEMA_VERSION,
    "base": "order",
    "columns": [{"field": "order.code"}],
    "filters": {
        "op": "and",
        "children": [
            {"field": "order.datetime", "operator": "relative_last_days", "value": 7}
        ],
    },
}


@pytest.mark.django_db
def test_relative_filter_is_evaluated_per_run_not_at_save_time(
    registered, event, user_with_perms
):
    """Two runs of one saved report under two frozen clocks, two results.

    The order is placed on 1 March. Run the "orders of the last 7 days" report
    on 3 March and it is in; run the very same stored report on 20 March and it
    must be gone. If anything resolved the window at save time -- in the form,
    in ``export_form_data``, in a cached compile -- both runs would agree and
    this test would fail.
    """
    with scopes_disabled():
        channel = event.organizer.sales_channels.get(identifier="web")
        Item.objects.create(event=event, name="Ticket", default_price=1)
        Order.objects.create(
            event=event,
            code="MARCH",
            status=Order.STATUS_PAID,
            email="m@example.org",
            sales_channel=channel,
            datetime=datetime.datetime(2026, 3, 1, 12, 0, tzinfo=datetime.timezone.utc),
            expires=datetime.datetime(2026, 4, 1, 12, 0, tzinfo=datetime.timezone.utc),
            total=Decimal("1.00"),
        )
        make_report(event=event, identifier="recent", definition=RELATIVE_DEFINITION)

    def run():
        with scope(organizer=event.organizer):
            ex = exporter_for_event(event, user_with_perms)
            return rows_of(ex, csv_form_data(identifier="recent"))

    with freeze_time("2026-03-03 09:00:00+00:00"):
        inside = run()
    with freeze_time("2026-03-20 09:00:00+00:00"):
        outside = run()

    assert [row[0] for row in inside[1:]] == ["MARCH"]
    assert outside[1:] == []


@pytest.mark.django_db
def test_a_scheduled_relative_report_re_evaluates_on_every_scheduled_run(
    registered, event, schedule_user
):
    """The same property through the scheduler, because that is where it matters.

    A daily "yesterday's orders" mail that silently keeps reporting the day the
    schedule was created is the kind of bug nobody notices for a quarter.
    """
    with scopes_disabled():
        channel = event.organizer.sales_channels.get(identifier="web")
        Item.objects.create(event=event, name="Ticket", default_price=1)
        Order.objects.create(
            event=event,
            code="MARCH",
            status=Order.STATUS_PAID,
            email="m@example.org",
            sales_channel=channel,
            datetime=datetime.datetime(2026, 3, 1, 12, 0, tzinfo=datetime.timezone.utc),
            expires=datetime.datetime(2026, 4, 1, 12, 0, tzinfo=datetime.timezone.utc),
            total=Decimal("1.00"),
        )
        make_report(event=event, identifier="recent", definition=RELATIVE_DEFINITION)
        schedule = make_schedule(event, schedule_user, identifier="recent")

    bodies = []
    for moment in ("2026-03-03 09:00:00+00:00", "2026-03-20 09:00:00+00:00"):
        with freeze_time(moment):
            djmail.outbox = []
            with scopes_disabled():
                schedule.schedule_next_run = now() - datetime.timedelta(minutes=5)
                schedule.save(update_fields=["schedule_next_run"])
            run_scheduled_exports(None)
            if djmail.outbox and djmail.outbox[0].attachments:
                payload = djmail.outbox[0].attachments[0][1]
                if isinstance(payload, bytes):
                    payload = payload.decode("utf-8")
                bodies.append(payload)
            else:
                bodies.append("")

    assert "MARCH" in bodies[0]
    assert "MARCH" not in bodies[1]


# ---------------------------------------------------------------------------
# 6. Injection: verify, do not duplicate
# ---------------------------------------------------------------------------


INJECTION = '=1+cmd|" /C calc"!A0'


@pytest.fixture
def injected_order(event):
    with scopes_disabled():
        channel = event.organizer.sales_channels.get(identifier="web")
        Order.objects.create(
            event=event,
            code="EVILL",
            status=Order.STATUS_PAID,
            email="e@example.org",
            sales_channel=channel,
            datetime=now(),
            expires=now() + datetime.timedelta(days=1),
            total=Decimal("1.00"),
            comment=INJECTION,
        )
        make_report(
            event=event,
            identifier="comments",
            definition={
                "schema_version": contracts.SCHEMA_VERSION,
                "base": "order",
                "columns": [{"field": "order.comment"}],
            },
        )


@pytest.mark.django_db
def test_csv_injection_is_neutralised_by_listexporter(
    registered, event, user_with_perms, injected_order
):
    """``ListExporter`` imports ``defusedcsv`` (pretix/base/exporter.py:42).

    We do not escape anything ourselves -- doing it twice would put two
    apostrophes in front of honest data. This test is the substitute for that
    code: it fails if a future refactor ever hand-rolls a ``csv.writer``.
    """
    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        _fn, _mime, content = ex.render(csv_form_data(identifier="comments"))
    body = content.decode("utf-8")
    assert INJECTION not in body
    assert "'=1+cmd" in body


@pytest.mark.django_db
def test_xlsx_injection_is_neutralised_by_safeworkbook(
    registered, event, user_with_perms, injected_order
):
    """``_render_xlsx`` builds a ``SafeWorkbook`` (pretix/base/exporter.py:298).

    ``SafeCell`` forces a cell openpyxl would type as a formula back to text
    (pretix/helpers/safe_openpyxl.py:75-84), so the value survives verbatim but
    Excel will not execute it.
    """
    import io
    from openpyxl import load_workbook

    with scope(organizer=event.organizer):
        ex = exporter_for_event(event, user_with_perms)
        _fn, _mime, content = render_xlsx(
            ex, {"_format": "xlsx", contracts.EXPORT_FORM_REPORT_KEY: "comments"}
        )
    sheet = load_workbook(io.BytesIO(content)).active
    cell = sheet.cell(row=2, column=1)
    assert cell.value == INJECTION
    assert cell.data_type == "s"  # string, not formula


# ---------------------------------------------------------------------------
# 7. Structural guarantees
# ---------------------------------------------------------------------------


def test_the_exporter_contains_no_query_logic_of_its_own():
    """CLAUDE.md rules 2 and 3, asserted rather than promised.

    Every row must come from the query compiler. Naming an ORM path, importing
    ``Order``/``OrderPosition`` or building a ``Q()`` here would move part of
    the allow-list out of the registry, which is the one place it is reviewed.
    """
    import pathlib

    source = (
        pathlib.Path(__file__).resolve().parent.parent
        / "pretix_custom_reports"
        / "exporters.py"
    ).read_text(encoding="utf-8")
    code = "\n".join(
        line for line in source.splitlines() if not line.strip().startswith("#")
    )
    # The module docstring mentions these names in prose; strip it first.
    code = code.split('"""', 2)[-1]
    for forbidden in (
        "eval(",
        "exec(",
        ".raw(",
        "RawSQL",
        ".extra(",
        "Q(",
        "OrderPosition",
        "filter(order__",
    ):
        assert forbidden not in code, forbidden


def test_the_registration_receivers_return_the_class_not_an_instance():
    """``init_*_exporters`` calls ``response(...)`` on the return value."""
    assert exporters.register_report_exporter(None) is exporters.CustomReportExporter
    assert (
        exporters.register_multievent_report_exporter(None)
        is exporters.CustomReportExporter
    )


def test_the_identifier_is_lowercase_and_short():
    """It becomes ``ScheduledExport.export_identifier`` and an HTML field prefix.

    ``max_length=190`` on the column, and the control UI renders form fields as
    ``"<identifier>-<fieldname>"`` (control/views/orders.py:2695-2699), so a
    dash or an uppercase letter here would be a lasting annoyance.
    """
    identifier = exporters.CustomReportExporter.identifier
    assert identifier.isalpha() and identifier.islower()
    assert len(identifier) <= 190


def test_signals_py_connects_both_receivers_at_plugin_import():
    """The wiring is production code's job, and this asserts it is done.

    Deliberately *not* using the ``registered`` fixture: the fixture would
    paper over a missing connection by establishing it itself. This reads the
    snapshot taken while the module was imported, i.e. the state
    ``apps.ready()`` -> ``signals.py`` left behind, and it names the two
    dispatch_uids literally because they are an interface -- they are what
    ``handoff/requests/exporter-dev-an-integrator-signals.md`` asked for and
    what every other test module disconnects by.
    """
    assert WIRING_AT_IMPORT[DISPATCH_UID] is exporters.register_report_exporter
    assert (
        WIRING_AT_IMPORT[MULTI_DISPATCH_UID]
        is exporters.register_multievent_report_exporter
    )


# NOTE: keep this last. pytest runs the tests of a module in definition order,
# so a check placed here has seen every ``registered`` teardown in this file.
def test_this_module_hands_the_signal_wiring_back_untouched(
    wiring_before_this_module,
):
    """The canary for the bug this file used to have.

    The old fixture connected and disconnected ``DISPATCH_UID`` around every
    test. Because ``disconnect(dispatch_uid=...)`` matches on the uid alone, the
    last teardown of this module removed the registration ``signals.py`` had
    made at plugin import -- for the rest of the pytest session. Nothing failed
    at the time, because whoever ran next either instantiated the exporter
    directly or, like tests/test_smoke.py, had already snapshotted the wiring.
    The next test that genuinely went through the export UI would have failed
    depending on file order, which is the worst possible way to find out.

    The first assertion is the one that belongs to this file: whatever we were
    handed, we hand back. The rest only runs when we really were handed the
    production wiring, so that this test keeps reporting *our* leaks and not
    another module's -- both tests/test_integration.py and tests/test_security.py
    still carry the old connect/disconnect pair, and either of them running
    first would otherwise turn this into a failure with the wrong name on it.
    """
    after = {
        dispatch_uid: connected_receiver(signal, dispatch_uid)
        for signal, _receiver, dispatch_uid in WIRING
    }
    assert after == wiring_before_this_module

    if wiring_before_this_module == WIRING_AT_IMPORT:
        # ``has_listeners`` and the identity check fail for different reasons:
        # a signal emptied wholesale versus our own uid dropped or rebound.
        assert register_data_exporters.has_listeners()
        assert register_multievent_data_exporters.has_listeners()
        for _signal, receiver, dispatch_uid in WIRING:
            assert after[dispatch_uid] is receiver, dispatch_uid
