# Owner from wave 3 on: test-engineer (see ORCHESTRIERUNG.md section 5)
"""End-to-end tests across every component, with hand-computed expectations.

Why this module exists
----------------------

Registry, query compiler, persistence, editor, exporter and portability were
written in parallel by different agents against a frozen contract. Each of them
has a thorough unit suite, and each of those suites is honest about what it
covers -- but a unit suite proves that a part works *against its own
assumptions*. The two most expensive bugs of the previous waves were both of the
other kind: the compiler read an aggregate condition from a convention only its
own test double implemented (``handoff/status/query-dev.md``, "Fehler 1"), and
the field library served the wrong of two ``ReportField`` objects for a choice
question (``handoff/status/frontend-dev.md``, wave 2). Neither produced an
exception. Both produced wrong numbers, and both were invisible until somebody
ran two components together.

So the rules for this module:

1. **Assertions are numbers, not shapes.** ``assert rows`` proves nothing;
   ``assert rows == [["PAID1", Decimal("33.00"), 2, "L, XL"]]`` proves something.
   Every expected value below was computed by hand from the ledger in
   ``tests/factories.py`` and is written out literally, never derived from the
   code under test or imported from the factory.
2. **The path is walked, not simulated.** The walk-through test goes through the
   editor's JSON API, the CRUD form, ``ListExporter``, the export view's file
   response and the import view's two-step confirmation -- over HTTP, through
   the real URL resolver and the real permission decorators -- rather than
   calling the four packages' functions in sequence.
3. **Failures point at one agent.** Each test name says what broke; the
   docstring says why it matters and, where relevant, what the wrong answer
   would look like.

Test data comes from :mod:`tests.factories`; the ledger it builds is documented
there and the expectations here were derived from that table, independently.
"""

from typing import Any, Dict, List

import csv
import datetime as dt
import io
import json
import openpyxl
import os
import pytest
import tempfile
import warnings
import weakref
from decimal import Decimal
from django.urls import reverse
from django_scopes import scope, scopes_disabled
from freezegun import freeze_time
from pretix.base.exporter import ListExporter
from pretix.base.models import Item, Order, Question, ScheduledEventExport
from pretix.base.services.export import (
    ExportError,
    init_event_exporter,
    run_scheduled_exports,
)
from pretix.base.signals import (
    register_data_exporters,
    register_multievent_data_exporters,
)

from pretix_custom_reports import contracts, exporters
from pretix_custom_reports.models import ReportDefinition
from pretix_custom_reports.portability.resolution import (
    STATUS_FOUND,
    STATUS_MAPPED,
    STATUS_MISSING,
    ResolutionStrategy,
)
from pretix_custom_reports.portability.templating import apply_template, plan_template
from pretix_custom_reports.query.compiler import ReportQueryCompiler
from pretix_custom_reports.registry import cache as registry_cache
from pretix_custom_reports.registry.library import field_registry

from . import factories

URL_NAMESPACE = "plugins:pretix_custom_reports"

#: The dispatch_uids ``signals.py`` uses (integrator, wave 4). Not a private
#: choice of this file any more: since the plugin connects both receivers at
#: import time, a test fixture that wants them connected has to talk about the
#: *same* key, otherwise it adds a second registration instead of guaranteeing
#: the first one. See :func:`registered`.
DISPATCH_UID = "pretix_custom_reports_exporter"
MULTI_DISPATCH_UID = "pretix_custom_reports_multiexporter"

#: The wiring these tests need, as ``(signal, receiver, dispatch_uid)`` -- the
#: same three-tuples ``signals.py`` passes to ``connect()``.
WIRING = (
    (register_data_exporters, exporters.register_report_exporter, DISPATCH_UID),
    (
        register_multievent_data_exporters,
        exporters.register_multievent_report_exporter,
        MULTI_DISPATCH_UID,
    ),
)


def connected_receiver(signal, dispatch_uid):
    """The receiver currently connected to *signal* under *dispatch_uid*.

    ``Signal.receivers`` holds ``(lookup_key, receiver, is_async)`` in Django
    5.2 and ``lookup_key`` is ``(dispatch_uid or id(receiver), id(sender))``
    (django/dispatch/dispatcher.py:96-99, 113-117). With the default
    ``weak=True`` the second slot is a ``weakref.ref`` and has to be
    dereferenced before it can be compared against a function.

    Returns ``None`` if nothing is connected under that dispatch_uid.

    Deliberately a copy of the identical helper in ``tests/test_exporters.py``
    rather than an import: the two modules belong to different agents
    (ORCHESTRIERUNG.md section 5) and one test module importing another's
    internals turns a rename over there into a failure over here.
    """
    for lookup_key, receiver, *_ in signal.receivers:
        if lookup_key[0] == dispatch_uid:
            if isinstance(receiver, weakref.ReferenceType):
                return receiver()
            return receiver
    return None


#: The production wiring as it stood while this module was imported, i.e. after
#: ``apps.ready()`` -> ``signals.py`` and before any test or fixture here ran.
#: ``signals.py`` is imported once per process and never re-connects itself, so
#: this is the only reliable record of what the plugin established at startup.
WIRING_AT_IMPORT = {
    dispatch_uid: connected_receiver(signal, dispatch_uid)
    for signal, _receiver, dispatch_uid in WIRING
}


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def clean_registry_cache():
    """The registry's process-local field cache is keyed by event primary key.

    Primary keys repeat across tests, and a ``ReportField`` holds closures over
    an event, so a leftover entry would serve one test's fields to the next.
    ``tests/test_query_registry.py`` does the same thing for the same reason.
    """
    registry_cache.clear_local_cache()
    yield
    registry_cache.clear_local_cache()


@pytest.fixture
def main_event(organizer):
    """The reference event. Deterministic ``date_from``, unlike the shared fixture.

    ``computed.age.*`` and ``relative_since_event_start`` both read
    ``Event.date_from``; with ``now() + 30 days`` the expected age would change
    on somebody's birthday.
    """
    return factories.make_event(organizer, slug="main", name="Main Event")


@pytest.fixture
def world(main_event):
    return factories.build_reference_world(main_event)


@pytest.fixture
def second_event(organizer):
    """A second event of the same organizer, plugin enabled, no data yet."""
    return factories.make_event(organizer, slug="second", name="Second Event")


@pytest.fixture(scope="module", autouse=True)
def wiring_before_this_module():
    """The signal wiring as this module found it, for the canary at the bottom.

    Module scoped and autouse so that it is established before the first test
    here runs. Not the same thing as :data:`WIRING_AT_IMPORT`: what this file
    has to answer for is that it changes nothing, not that the session was
    healthy when it got the process.
    """
    return {
        dispatch_uid: connected_receiver(signal, dispatch_uid)
        for signal, _receiver, dispatch_uid in WIRING
    }


@pytest.fixture
def registered():
    """Guarantee the two exporter receivers are connected -- and restore after.

    Since wave 4 ``signals.py`` connects both receivers at plugin import
    (``apps.ready()``), so in a normal run this fixture has nothing left to do.
    It must nevertheless not be written as a plain connect/disconnect pair, and
    the reason is not style:

    * ``Signal.connect()`` skips a receiver whose ``(dispatch_uid, sender_id)``
      key is already present (django/dispatch/dispatcher.py:113-117), and
    * ``Signal.disconnect(dispatch_uid=...)`` matches on that key *alone* --
      the receiver argument is ignored entirely (dispatcher.py:138-153).

    Neither call knows who connected first. So a fixture using the production
    uids would connect nothing and, on teardown, remove the *production*
    registration for the rest of the pytest session, in whatever file follows.
    pretix' ``EventPluginSignal``/``OrganizerPluginSignal`` override ``connect``
    but not ``disconnect`` (pretix/base/signals.py:261-311), so nothing softens
    this for plugin signals. That is the defect ``exporter-dev`` found in
    ``tests/test_exporters.py``; ``tests/test_smoke.py`` still carries a
    snapshot workaround because of it.

    This module dodged that particular failure by using private, ``_integration``
    suffixed uids -- and bought a quieter one for it: two distinct keys pointing
    at the same function are two receivers, so ``register_data_exporters.send()``
    returned :class:`CustomReportExporter` twice and every export list in this
    file silently contained our exporter twice over.
    ``init_event_exporter()`` returns the first match
    (pretix/base/services/export.py:191-195), which is why nothing ever failed;
    the export *page* would have listed the report export twice.
    ``test_the_event_export_is_offered_exactly_once`` is the guard for that.

    So: use the production uids, connect only what is missing, disconnect only
    what we connected. If a uid is already taken, assert it is taken by the
    function we expect -- otherwise the no-op ``connect()`` would quietly run
    this whole module against somebody else's receiver.

    ``register_multievent_data_exporters`` is an
    ``OrganizerPluginSignal(allow_legacy_plugins=True)`` and this plugin is
    event level, so ``connect()`` emits a ``DeprecationWarning``
    (pretix/base/signals.py:301-306) that pretix filters in its own test config
    and we do not. Silenced narrowly, so a *different* deprecation stays visible.
    """
    connected_by_us = []
    for signal, receiver, dispatch_uid in WIRING:
        existing = connected_receiver(signal, dispatch_uid)
        if existing is not None:
            assert existing is receiver, (
                f"{dispatch_uid!r} is connected to {existing!r}, not to "
                f"{receiver!r} -- signals.py and exporters.py disagree."
            )
            continue
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message=".*organizer-level.*",
                category=DeprecationWarning,
            )
            signal.connect(receiver, dispatch_uid=dispatch_uid)
        connected_by_us.append((signal, dispatch_uid))
    yield
    for signal, dispatch_uid in connected_by_us:
        signal.disconnect(dispatch_uid=dispatch_uid)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def definition(base: str = "order", **parts: Any) -> Dict[str, Any]:
    """A definition document as the editor would post it."""
    document: Dict[str, Any] = {
        "schema_version": contracts.SCHEMA_VERSION,
        "base": base,
        "columns": [],
    }
    document.update(parts)
    return document


def columns(*keys: Any) -> List[Dict[str, Any]]:
    """``columns`` from ``"key"`` or ``("key", "aggregate")`` pairs."""
    built = []
    for key in keys:
        if isinstance(key, tuple):
            built.append({"field": key[0], "aggregate": key[1]})
        else:
            built.append({"field": key})
    return built


def run_report(document: Dict[str, Any], event: Any, **kwargs: Any):
    """Compile and execute *document* for *event*; return ``(headers, rows)``."""
    parsed = contracts.validate_definition(document)
    compiler = ReportQueryCompiler(field_registry())
    with scopes_disabled():
        report = compiler.compile(parsed, event, **kwargs)
        return report.headers(), list(report.iter_rows())


def rows_by_code(rows: List[List[Any]]) -> Dict[Any, List[Any]]:
    """Index rows by their first cell. Every report here starts with a code."""
    return {row[0]: row for row in rows}


def url_for(name: str, event: Any, **kwargs: Any) -> str:
    return reverse(
        f"{URL_NAMESPACE}:{name}",
        kwargs={"organizer": event.organizer.slug, "event": event.slug, **kwargs},
    )


def org_url(name: str, organizer: Any, **kwargs: Any) -> str:
    return reverse(
        f"{URL_NAMESPACE}:{name}", kwargs={"organizer": organizer.slug, **kwargs}
    )


def post_json(client, url: str, payload: Dict[str, Any]):
    return client.post(url, data=json.dumps(payload), content_type="application/json")


def store_report(
    event: Any,
    document: Dict[str, Any],
    *,
    identifier: str = "report",
    name: str = "Report",
) -> ReportDefinition:
    with scopes_disabled():
        return ReportDefinition.objects.create(
            event=event, name=name, identifier=identifier, definition=document
        )


def export_csv(event: Any, user: Any, identifier: str, **form_extra: Any):
    """Run the report through the real exporter and parse the CSV back.

    Goes through ``init_event_exporter``, which is what the export page and the
    scheduler both use, so the permission check and the form-data path are part
    of the test rather than bypassed.
    """
    with scope(organizer=event.organizer):
        exporter = init_event_exporter(
            identifier=exporters.CustomReportExporter.identifier,
            event=event,
            user=user,
        )
        form_data = {
            "_format": "default",
            contracts.EXPORT_FORM_REPORT_KEY: identifier,
            **form_extra,
        }
        filename, content_type, payload = exporter.render(form_data)
    text = payload.decode("utf-8-sig")
    return filename, list(csv.reader(io.StringIO(text)))


def export_rows(exporter, form_data):
    """``iterate_list`` without the progress marker."""
    return [
        line
        for line in exporter.iterate_list(form_data)
        if not isinstance(line, ListExporter.ProgressSetTotal)
    ]


def export_xlsx(event: Any, user: Any, identifier: str, **form_extra: Any):
    """Run the report as XLSX and read the sheet back. Returns a list of rows.

    Written into a file handle we opened ourselves, deliberately.
    ``ListExporter._render_xlsx`` *without* ``output_file`` saves into a
    ``NamedTemporaryFile`` and then opens it a second time by name, which raises
    ``PermissionError`` on Windows -- a platform limit of pretix, not of this
    plugin (``handoff/status/exporter-dev.md``). Passing ``output_file`` is the
    path pretix' own export service uses for anything that goes to a file, so
    this is the realistic call and not a workaround.
    """
    with scope(organizer=event.organizer):
        exporter = init_event_exporter(
            identifier=exporters.CustomReportExporter.identifier,
            event=event,
            user=user,
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
            path = handle.name
        with open(path, "wb") as handle:
            exporter.render(
                {
                    "_format": "xlsx",
                    contracts.EXPORT_FORM_REPORT_KEY: identifier,
                    **form_extra,
                },
                output_file=handle,
            )
    book = openpyxl.load_workbook(path)
    try:
        return [list(row) for row in book.active.values]
    finally:
        book.close()
        os.unlink(path)


# ===========================================================================
# 1. The walk-through: editor -> save -> run -> export -> file -> other event
# ===========================================================================


@pytest.mark.django_db
def test_the_whole_path_from_the_editor_to_another_events_export(
    wired_urls, registered, client_with_perms, user_with_perms, world, second_event
):
    """One report, all the way round, checked at every station.

    Editor page -> ``api/validate`` -> ``api/preview`` -> CRUD save -> CSV export
    -> JSON file -> import into a second event -> CSV export there.

    The second event gets its **own** data and its **own** questions with the
    same identifiers, which is the realistic case for a template: the definition
    has to travel, the rows must not. The last assertion is therefore not "the
    export ran" but "the export contains the second event's two orders and
    exactly their numbers".
    """
    main = world.event

    # -- the second event, with the same question identifiers ---------------
    with scopes_disabled():
        other_catalog = factories.make_catalog(second_event)
        other_questions = factories.make_questions(second_event)
        other_paid = factories.make_order(
            second_event, "ZZZ11", Order.STATUS_PAID, Decimal("46.00")
        )
        z1 = factories.add_position(
            other_paid, other_catalog.ticket, Decimal("23.00"), 1
        )
        factories.add_position(other_paid, other_catalog.ticket, Decimal("23.00"), 2)
        factories.add_payment(other_paid, Decimal("46.00"))
        factories.add_answer(
            z1,
            other_questions["tshirt-size"],
            "M",
            [other_questions["tshirt-size"].options.get(answer="M")],
        )
        other_pending = factories.make_order(
            second_event, "ZZZ22", Order.STATUS_PENDING, Decimal("10.00")
        )
        factories.add_position(
            other_pending, other_catalog.workshop, Decimal("10.00"), 1
        )

    document = definition(
        base="order",
        columns=columns(
            "order.code",
            "order.status",
            "payment.sum_confirmed",
            ("answer.tshirt-size", "join"),
        ),
        sorting=[{"field": "order.code", "direction": "asc"}],
    )

    # -- 1. the editor opens ------------------------------------------------
    response = client_with_perms.get(url_for("editor.new", main))
    assert response.status_code == 200

    # -- 2. the draft validates and comes back canonical --------------------
    response = post_json(
        client_with_perms, url_for("api.validate", main), {"definition": document}
    )
    assert response.status_code == 200
    validated = response.json()
    assert validated["ok"] is True, validated
    assert validated["warnings"] == []
    canonical = validated["definition"]

    # -- 3. the preview shows this event's data -----------------------------
    response = post_json(
        client_with_perms, url_for("api.preview", main), {"definition": canonical}
    )
    assert response.status_code == 200
    preview = response.json()
    assert [row[0] for row in preview["rows"]] == [
        "CANC5",
        "EXPI4",
        "OVER6",
        "PAID1",
        "PART2",
        "PEND3",
    ]

    # -- 4. saving through the CRUD form ------------------------------------
    response = client_with_perms.post(
        url_for("event.reports.add", main),
        data={
            "name": "Payment overview",
            "description": "",
            "identifier": "",
            "base": "order",
            "definition": json.dumps(canonical),
        },
    )
    assert response.status_code == 302, getattr(response, "context_data", None)
    with scopes_disabled():
        stored = ReportDefinition.objects.get(event=main, name="Payment overview")
    assert stored.definition == canonical
    identifier = stored.identifier

    # -- 5. the exporter produces exactly these lines -----------------------
    filename, lines = export_csv(main, user_with_perms, identifier)
    assert filename.endswith(".csv")
    assert lines[0] == ["Order code", "Order status", "Amount paid", "T-shirt size"]
    # The money cell is compared as a number rather than as a string. It is
    # written by the database driver and its *scale* differs between backends
    # for an aggregate column -- SQLite writes "23", PostgreSQL "23.00". That is
    # a real defect, but it is not this test's subject; see
    # ``test_an_aggregated_money_column_keeps_its_two_decimal_places`` and
    # handoff/blockers.md finding 2.
    assert [[row[0], row[1], Decimal(row[2]), row[3]] for row in lines[1:]] == [
        ["CANC5", "c", Decimal("23.00"), ""],
        ["EXPI4", "e", Decimal("0.00"), ""],
        ["OVER6", "p", Decimal("30.00"), ""],
        ["PAID1", "p", Decimal("33.00"), "L, XL"],
        ["PART2", "n", Decimal("20.00"), "M"],
        ["PEND3", "n", Decimal("0.00"), ""],
    ]

    # -- 6. exporting the definition as a file ------------------------------
    response = client_with_perms.get(
        url_for("event.reports.export", main, report=stored.pk)
    )
    assert response.status_code == 200
    assert response["Content-Type"] == "application/json"
    file_text = response.content.decode("utf-8")
    portable = json.loads(file_text)
    contracts.validate_portable_document(portable)
    assert portable["definition"] == canonical
    # What makes the file portable: slugs and identifiers, no primary keys.
    assert portable["source"] == "dummy/main"
    assert portable["meta"]["identifier"] == identifier
    assert {
        reference["identifier"] for reference in portable["meta"]["references"]
    } == {"tshirt-size"}

    # -- 7. importing it into the second event, two steps --------------------
    import_url = url_for("event.reports.import", second_event)
    response = client_with_perms.post(import_url, data={"text": file_text})
    assert response.status_code == 200
    assert response.context["plan"].ok is True

    response = client_with_perms.post(
        import_url, data={"document": file_text, "action": "confirm"}
    )
    assert response.status_code == 302
    with scopes_disabled():
        imported = ReportDefinition.objects.get(event=second_event)
    assert imported.definition == canonical
    assert imported.identifier == identifier, (
        "the identifier travels with the report, otherwise every scheduled "
        "export referencing it breaks (ADR 0001 section 5.1)"
    )

    # -- 8. and it produces the second event's numbers ----------------------
    filename, lines = export_csv(second_event, user_with_perms, identifier)
    assert lines[0] == ["Order code", "Order status", "Amount paid", "T-shirt size"]
    assert [[row[0], row[1], Decimal(row[2]), row[3]] for row in lines[1:]] == [
        ["ZZZ11", "p", Decimal("46.00"), "M"],
        ["ZZZ22", "n", Decimal("0.00"), ""],
    ]


@pytest.mark.django_db
def test_a_template_reaches_an_event_whose_question_is_spelled_differently(
    wired_urls, registered, client_with_perms, user_with_perms, world, organizer
):
    """Organizer template -> event, with one key that only matches by spelling.

    The target event calls the question ``tshirt_size`` (underscore) and has no
    ``diet`` question at all. Expected: one key found, one mapped, one missing,
    and the ``skip`` strategy produces a *reduced* report that runs -- with the
    dropped column really gone from the CSV, not silently empty.
    """
    with scopes_disabled():
        source = ReportDefinition.objects.create(
            event=world.event,
            name="Sizes",
            identifier="sizes",
            definition=definition(
                base="orderposition",
                columns=columns("position.code", "answer.tshirt-size", "answer.diet"),
                sorting=[{"field": "order.code", "direction": "asc"}],
            ),
        )
        template = source.duplicate(event=None, organizer=organizer)

        target = factories.make_event(organizer, slug="target", name="Target")
        target_catalog = factories.make_catalog(target)
        Question.objects.create(
            event=target,
            question="T-shirt size",
            identifier="tshirt_size",
            type=Question.TYPE_CHOICE,
            position=0,
        )
        order = factories.make_order(
            target, "TGT01", Order.STATUS_PAID, Decimal("23.00")
        )
        position = factories.add_position(
            order, target_catalog.ticket, Decimal("23.00"), 1
        )
        factories.add_answer(
            position, Question.objects.get(event=target, identifier="tshirt_size"), "S"
        )

    with scopes_disabled():
        plan = plan_template(template, target, user=user_with_perms)
        statuses = {entry.source: entry.status for entry in plan.report.fields}
    assert statuses["position.code"] == STATUS_FOUND
    assert statuses["answer.tshirt-size"] == STATUS_MAPPED
    assert statuses["answer.diet"] == STATUS_MISSING
    assert plan.ok is False, "abort is the default and one key is missing"

    with scopes_disabled():
        skipping = plan_template(
            template, target, strategy=ResolutionStrategy.SKIP, user=user_with_perms
        )
        assert skipping.ok is True
        copy = apply_template(skipping, user=user_with_perms)

    assert [column["field"] for column in copy.definition["columns"]] == [
        "position.code",
        "answer.tshirt_size",
    ]
    _, lines = export_csv(target, user_with_perms, copy.identifier)
    assert lines == [["Position code", "T-shirt size"], ["TGT01-1", "S"]]


@pytest.mark.django_db
def test_an_event_copy_carries_its_reports_and_runs_them_in_the_copy(
    registered, user_with_perms, world, organizer
):
    """``Event.copy_data_from`` alone must bring the reports along and run them.

    Nothing here calls ``copy_reports_to_event``. Since wave 4 the receiver in
    ``signals.py`` is connected to ``event_copy_data``, so the copy happens as a
    side effect of the pretix call -- which is the only way to find out whether
    the receiver is wired at all, whether the signal reaches an event-level
    plugin, and whether the copy sees the *new* event's questions. The unit
    tests in ``tests/test_portability.py`` section 8 cover the function itself.

    ``copy_reports_to_event`` resolves with ``KEEP``, so a key the target does
    not have stays in the definition rather than disappearing. This checks the
    happy path end to end: the copied event has the same questions (because
    pretix copies questions with their identifiers), so the copied report runs
    and returns the *copy's* orders, not the source's.
    """
    with scopes_disabled():
        source_report = ReportDefinition.objects.create(
            event=world.event,
            name="Sizes",
            identifier="sizes",
            definition=definition(
                base="orderposition",
                columns=columns("order.code", "answer.tshirt-size"),
                sorting=[{"field": "order.code", "direction": "asc"}],
            ),
        )
        copy_event = factories.make_event(organizer, slug="copy", name="Copy")
        copy_event.copy_data_from(world.event)

        # One report in, exactly one report out, under its own identifier. A
        # receiver that fires twice would leave a second row called "sizes-2"
        # (``ReportDefinition`` deduplicates identifiers per event), and that is
        # the failure this list comparison is here to catch.
        copies = list(copy_event.custom_reports.order_by("identifier"))

    assert [report.identifier for report in copies] == ["sizes"]
    copied = copies[0]
    assert copied.pk != source_report.pk, "the copy must be a new row"
    assert copied.definition == source_report.definition

    # A copied event has the products and questions but no orders.
    _, lines = export_csv(copy_event, user_with_perms, copied.identifier)
    assert lines == [["Order code", "T-shirt size"]]

    with scopes_disabled():
        item = Item.objects.get(event=copy_event, internal_name="ticket")
        question = Question.objects.get(event=copy_event, identifier="tshirt-size")
        order = factories.make_order(
            copy_event, "CPY01", Order.STATUS_PAID, Decimal("23.00")
        )
        position = factories.add_position(order, item, Decimal("23.00"), 1)
        factories.add_answer(position, question, "XL")

    _, lines = export_csv(copy_event, user_with_perms, copied.identifier)
    assert lines == [["Order code", "T-shirt size"], ["CPY01", "XL"]]


@pytest.mark.django_db
def test_the_event_copy_signal_hands_the_question_map_to_the_log_entry(
    world, organizer
):
    """The receiver's arguments arrive intact -- checked through the log entry.

    ``event_copy_data`` sends seven ``*_map`` arguments and the receiver takes
    ``question_map`` out of them; the copy writes what it got into the
    ``copied_from_event`` block of its ``report.added`` entry. That block is the
    only place where the *arguments of the real signal* become observable, so
    this is what separates "the receiver is connected" from "the receiver is
    connected and pretix passes it what we think it does".

    Expected values, all hand-counted from ``tests/factories.py``:

    * ``questions_mapped == 12`` -- ``QUESTION_SPECS`` has one entry per
      ``Question.type`` and ``build_reference_world`` creates all of them, and
      pretix' event copy takes every question along,
    * source organizer/event/identifier as written below,
    * ``resolution`` clean, because the copy has the same question identifiers:
      both column references resolve exactly, nothing is mapped or missing.

    The entry has no user on purpose: ``event_copy_data`` does not carry one,
    so the receiver cannot invent one. The copy is attributed to the event copy
    the user triggered, which pretix logs separately on the event.
    """
    with scopes_disabled():
        ReportDefinition.objects.create(
            event=world.event,
            name="Sizes",
            identifier="sizes",
            definition=definition(
                base="orderposition",
                columns=columns("order.code", "answer.tshirt-size"),
            ),
        )
        copy_event = factories.make_event(organizer, slug="copy", name="Copy")
        copy_event.copy_data_from(world.event)

        copied = copy_event.custom_reports.get(identifier="sizes")
        entries = list(copied.all_logentries())
        assert [entry.action_type for entry in entries] == [contracts.LOG_ACTION_ADDED]
        entry = entries[0]
        assert entry.user is None
        assert entry.event_id == copy_event.pk
        provenance = entry.parsed_data["copied_from_event"]

    assert provenance["organizer"] == organizer.slug
    assert provenance["event"] == world.event.slug
    assert provenance["identifier"] == "sizes"
    assert provenance["questions_mapped"] == 12
    assert provenance["resolution"]["issues"] == []
    counts = provenance["resolution"]["counts"]
    assert counts["found"] == 2, "order.code and answer.tshirt-size"
    assert counts["mapped"] == 0
    assert counts["missing"] == 0


@pytest.mark.django_db
def test_a_multi_event_export_labels_every_row_with_its_event(
    registered, user_with_perms, world, organizer
):
    """The organizer-level export over two events that hold the same identifier.

    Two leading columns, one compile per event, and -- the part that is easy to
    get wrong -- the rows of the two events must not be interleaved or
    duplicated. Six orders here, one there.
    """
    with scopes_disabled():
        ReportDefinition.objects.create(
            event=world.event,
            name="Codes",
            identifier="codes",
            definition=definition(
                base="order",
                columns=columns("order.code"),
                sorting=[{"field": "order.code", "direction": "asc"}],
            ),
        )
        other = factories.make_event(organizer, slug="other", name="Other Event")
        catalog = factories.make_catalog(other)
        order = factories.make_order(
            other, "OTH01", Order.STATUS_PAID, Decimal("23.00")
        )
        factories.add_position(order, catalog.ticket, Decimal("23.00"), 1)
        ReportDefinition.objects.create(
            event=other,
            name="Codes",
            identifier="codes",
            definition=definition(
                base="order",
                columns=columns("order.code"),
                sorting=[{"field": "order.code", "direction": "asc"}],
            ),
        )

    from pretix.base.services.export import init_organizer_exporters

    with scope(organizer=organizer):
        exporter = next(
            ex
            for ex in init_organizer_exporters(
                organizer=organizer, user=user_with_perms
            )
            if ex.identifier == exporters.CustomReportExporter.identifier
        )
        rows = export_rows(
            exporter,
            {"_format": "default", contracts.EXPORT_FORM_REPORT_KEY: "codes"},
        )

    assert rows[0] == ["Event slug", "Event name", "Order code"]
    assert rows[1:] == [
        ["main", "Main Event", "CANC5"],
        ["main", "Main Event", "EXPI4"],
        ["main", "Main Event", "OVER6"],
        ["main", "Main Event", "PAID1"],
        ["main", "Main Event", "PART2"],
        ["main", "Main Event", "PEND3"],
        ["other", "Other Event", "OTH01"],
    ]


# ===========================================================================
# 2. Correctness: hand-computed numbers
# ===========================================================================


@pytest.mark.django_db
def test_the_money_columns_of_an_order_report_match_the_ledger(world):
    """Every money and counting column of the reference event, in one table.

    Recomputed by hand from ``tests/factories.py``:

    ===== ====== ===== ======== ======= ============== ==== =====
    code  total  paid  refunded pending state          pos. scans
    ===== ====== ===== ======== ======= ============== ==== =====
    PAID1 33.00  33.00 0.00     0.00    paid           2    2
    PART2 46.00  20.00 0.00     26.00   partially_paid 2    0
    PEND3 23.00  0.00  0.00     23.00   unpaid         1    0
    EXPI4 15.00  0.00  0.00     15.00   unpaid         1    0
    CANC5 23.00  23.00 23.00    0.00    paid           0    0
    OVER6 23.00  30.00 0.00     -7.00   overpaid       1    0
    ===== ====== ===== ======== ======= ============== ==== =====

    The three interesting cells: PART2 paid 25.00 in total but only 20.00 of it
    is in a state pretix counts, CANC5 is canceled so it owes nothing even though
    23.00 was paid for it, and PAID1 has four check-in rows of which only two are
    successful entry scans.
    """
    document = definition(
        base="order",
        columns=columns(
            "order.code",
            "order.total",
            "payment.sum_confirmed",
            "refund.sum_done",
            "order.pending_sum",
            "computed.payment_state",
            "order.position_count",
            "checkin.count",
        ),
        sorting=[{"field": "order.code", "direction": "asc"}],
    )
    _, rows = run_report(document, world.event)
    assert rows == [
        [
            "CANC5",
            Decimal("23.00"),
            Decimal("23.00"),
            Decimal("23.00"),
            Decimal("0.00"),
            "paid",
            0,
            0,
        ],
        [
            "EXPI4",
            Decimal("15.00"),
            Decimal("0.00"),
            Decimal("0.00"),
            Decimal("15.00"),
            "unpaid",
            1,
            0,
        ],
        [
            "OVER6",
            Decimal("23.00"),
            Decimal("30.00"),
            Decimal("0.00"),
            Decimal("-7.00"),
            "overpaid",
            1,
            0,
        ],
        [
            "PAID1",
            Decimal("33.00"),
            Decimal("33.00"),
            Decimal("0.00"),
            Decimal("0.00"),
            "paid",
            2,
            2,
        ],
        [
            "PART2",
            Decimal("46.00"),
            Decimal("20.00"),
            Decimal("0.00"),
            Decimal("26.00"),
            "partially_paid",
            2,
            0,
        ],
        [
            "PEND3",
            Decimal("23.00"),
            Decimal("0.00"),
            Decimal("0.00"),
            Decimal("23.00"),
            "unpaid",
            1,
            0,
        ],
    ]


@pytest.mark.django_db
def test_four_aggregates_over_three_relations_do_not_multiply_each_other(world):
    """The cross-product trap, with real data and four aggregates at once.

    ``positions``, ``answers`` (two hops through positions) and ``payments`` are
    three separate one-to-many relations of an order. Expressed as joined
    ``annotate(Sum(...))`` calls, PAID1's position sum would be counted once per
    matching answer and vice versa -- 33.00 would become 66.00 and the answer
    count 2 would become 4. ``query/relations.py`` uses correlated subqueries
    instead (``Order.annotate_overpayments`` is the precedent); this is that
    decision measured rather than described.

    PAID1 by hand: live positions 23.00 + 10.00 = 33.00, two of them; live
    T-shirt answers ``L`` and ``XL``; one counted payment of 33.00.
    """
    document = definition(
        base="order",
        columns=columns(
            "order.code",
            ("position.price", "sum"),
            ("position.positionid", "count"),
            ("answer.tshirt-size", "count"),
            ("answer.tshirt-size", "join"),
            "payment.sum_confirmed",
        ),
        sorting=[{"field": "order.code", "direction": "asc"}],
    )
    _, rows = run_report(document, world.event)
    indexed = rows_by_code(rows)
    assert indexed["PAID1"] == [
        "PAID1",
        Decimal("33.00"),
        2,
        2,
        "L, XL",
        Decimal("33.00"),
    ]
    assert indexed["PART2"] == [
        "PART2",
        Decimal("46.00"),
        2,
        1,
        "M",
        Decimal("20.00"),
    ]
    assert len(rows) == 6, "one row per order, never one row per position"


@pytest.mark.django_db
def test_the_naive_join_really_does_produce_the_wrong_number(world):
    """The counter-example, so that the test above is not vacuous.

    A green test proves nothing if the thing it guards against cannot happen. So
    here is the obvious implementation -- two ``annotate()`` aggregates over the
    same joined path -- run against the same data. PAID1 has two live positions
    (23.00 and 10.00) and thirteen live answers, so the join produces thirteen
    rows and the position sum comes out as 23.00 x 12 + 10.00 = 286.00 instead of
    33.00.

    This asserts on Django's behaviour, not on ours, which is unusual and
    deliberate: it is the reason ``query/relations.py`` uses correlated
    subqueries, and if a future Django ever stopped doing this, the reasoning in
    that module would deserve a re-read.
    """
    from django.db.models import Count, Q, Sum
    from pretix.base.models import Order as OrderModel

    with scopes_disabled():
        naive = (
            OrderModel.objects.filter(event=world.event, code="PAID1")
            .annotate(
                position_sum=Sum(
                    "all_positions__price",
                    filter=Q(all_positions__canceled=False),
                ),
                answer_count=Count(
                    "all_positions__answers",
                    filter=Q(all_positions__canceled=False),
                ),
            )
            .get()
        )
    assert naive.position_sum != Decimal("33.00"), (
        "if this ever equals 33.00, Django stopped multiplying rows across a "
        "join and query/relations.py can be simplified"
    )
    assert naive.position_sum == Decimal("286.00")
    assert naive.answer_count == 13


@pytest.mark.django_db
def test_two_answer_columns_keep_their_own_questions_apart(world):
    """Twelve questions on the event, two of them in the report.

    Without the question condition from ``registry.hints`` every answer cell
    would contain every answer of the order. PAID1's live positions hold twelve
    answers for position 1 plus one for position 2, so the T-shirt cell would
    read like a diary. Nothing would raise -- that is the point.
    """
    document = definition(
        base="order",
        columns=columns(
            "order.code",
            ("answer.tshirt-size", "join"),
            ("answer.tshirt-size", "count"),
            ("answer.nickname", "join"),
            ("answer.newsletter", "join"),
        ),
    )
    _, rows = run_report(document, world.event)
    assert rows_by_code(rows)["PAID1"][1:] == ["L, XL", 2, "Ada", "True"]


@pytest.mark.django_db
def test_include_canceled_positions_changes_exactly_four_cells(world):
    """One option, four code paths, and a defined effect on each.

    PAID1 has a canceled 10.00 position that answered the T-shirt question with
    ``S``; CANC5's only position is canceled. Turning the option on must move
    PAID1's position sum 33.00 -> 43.00, its count 2 -> 3, its joined answers
    ``L, XL`` -> ``L, XL, S``, and give CANC5 a sum at all. ``order.position_count``
    is *not* affected: the registry documents it as counting live positions only,
    so a change there would be a contradiction between two columns of one report.
    """
    document = definition(
        base="order",
        columns=columns(
            "order.code",
            ("position.price", "sum"),
            ("position.positionid", "count"),
            ("answer.tshirt-size", "join"),
            "order.position_count",
        ),
        options={"include_canceled_positions": False},
    )
    _, without = run_report(document, world.event)
    assert rows_by_code(without)["PAID1"][1:] == [
        Decimal("33.00"),
        2,
        "L, XL",
        2,
    ]
    assert rows_by_code(without)["CANC5"][1:] == [None, 0, "", 0]

    document["options"] = {"include_canceled_positions": True}
    _, with_canceled = run_report(document, world.event)
    assert rows_by_code(with_canceled)["PAID1"][1:] == [
        Decimal("43.00"),
        3,
        "L, XL, S",
        2,
    ]
    assert rows_by_code(with_canceled)["CANC5"][1:] == [
        Decimal("23.00"),
        1,
        "",
        0,
    ]


@pytest.mark.django_db
def test_include_testmode_orders_adds_exactly_the_test_order(world):
    """Off by default. The reference event has exactly one test-mode order."""
    document = definition(base="order", columns=columns("order.code"))
    _, default_rows = run_report(document, world.event)
    assert sorted(row[0] for row in default_rows) == [
        "CANC5",
        "EXPI4",
        "OVER6",
        "PAID1",
        "PART2",
        "PEND3",
    ]

    document["options"] = {"include_testmode_orders": True}
    _, with_test = run_report(document, world.event)
    assert sorted(row[0] for row in with_test) == [
        "CANC5",
        "EXPI4",
        "OVER6",
        "PAID1",
        "PART2",
        "PEND3",
        "TEST7",
    ]


@pytest.mark.django_db
def test_a_position_report_renders_an_answer_of_every_question_type(world):
    """All twelve ``Question.type`` values in one row, as the export writes them.

    ``QuestionAnswer.answer`` is a ``TextField`` for every type, so what a cell
    contains is decided by what pretix stored, not by the datatype the registry
    assigns. This nails the actual strings down: a change in the rendering of
    booleans or multiple-choice answers has to be a decision, not a side effect.
    """
    document = definition(
        base="orderposition",
        columns=columns(
            "position.code",
            "answer.tshirt-size",
            "answer.diet",
            "answer.nickname",
            "answer.notes",
            "answer.companions",
            "answer.newsletter",
            "answer.birthdate",
            "answer.arrival-time",
            "answer.arrival",
            "answer.home-country",
            "answer.phone",
            "answer.passport",
        ),
        filters={
            "op": "and",
            "children": [
                {"field": "order.code", "operator": "exact", "value": "PAID1"},
                {"field": "position.positionid", "operator": "exact", "value": 1},
            ],
        },
    )
    _, rows = run_report(document, world.event)
    assert rows == [
        [
            "PAID1-1",
            "L",
            "Vegan, Gluten-free",
            "Ada",
            "Two lines\nof text",
            "2",
            True,
            "1990-06-15",
            "14:30:00",
            "2026-06-03T14:30:00+02:00",
            "GB",
            "+441234567890",
            "file://passport.pdf",
        ]
    ]


@pytest.mark.django_db
def test_age_at_the_event_date_is_computed_in_the_database(world):
    """``computed.age.birthdate`` against the event's ``date_from``.

    The event starts on 2026-06-03. Ada was born 1990-06-15, so on the day the
    event starts she is **35**, not 36 -- her birthday is twelve days away. The
    second answer, 2010-06-15, gives 15 for the same reason. An implementation
    that subtracts years without looking at month and day says 36 and 16.
    """
    assert world.event.date_from.date() == dt.date(2026, 6, 3)
    document = definition(
        base="orderposition",
        columns=columns("position.code", "computed.age.birthdate"),
        sorting=[
            {"field": "order.code", "direction": "asc"},
            {"field": "position.positionid", "direction": "asc"},
        ],
    )
    _, rows = run_report(document, world.event)
    ages = {row[0]: row[1] for row in rows}
    assert ages["PAID1-1"] == 35
    assert ages["PART2-1"] == 15
    assert ages["PEND3-1"] is None


@pytest.mark.django_db
def test_filters_sorting_and_a_column_format_produce_this_exact_table(world):
    """A report that uses every part of a definition at once.

    Filter: orders with an outstanding amount over 10.00 **or** an overpayment.
    Sort: outstanding amount descending, then code ascending. Format: the
    ``join`` separator, which is the one format option the compiler does apply
    (see the finding above for the ones it does not).

    By hand -- outstanding amounts are 0.00 (PAID1), 26.00 (PART2), 23.00
    (PEND3), 15.00 (EXPI4), 0.00 (CANC5), -7.00 (OVER6). Over 10.00: PART2,
    PEND3, EXPI4. Below zero: OVER6. Descending: 26.00, 23.00, 15.00, -7.00.
    The joined product names follow the positions: PART2 has two tickets, the
    other three one each.
    """
    document = definition(
        base="order",
        columns=[
            {"field": "order.code"},
            {"field": "order.pending_sum"},
            {
                "field": "item.internal_name",
                "aggregate": "join",
                "format": {"separator": " | "},
            },
        ],
        filters={
            "op": "or",
            "children": [
                {"field": "order.pending_sum", "operator": "gt", "value": "10.00"},
                {"field": "order.pending_sum", "operator": "lt", "value": "0.00"},
            ],
        },
        sorting=[
            {"field": "order.pending_sum", "direction": "desc"},
            {"field": "order.code", "direction": "asc"},
        ],
    )
    _, rows = run_report(document, world.event)
    assert rows == [
        ["PART2", Decimal("26.00"), "ticket | ticket"],
        ["PEND3", Decimal("23.00"), "ticket"],
        ["EXPI4", Decimal("15.00"), "merch"],
        ["OVER6", Decimal("-7.00"), "ticket"],
    ]


@pytest.mark.django_db
def test_a_hidden_column_filters_and_sorts_but_does_not_appear(world):
    """``hidden`` is an output property, not a query property.

    A hidden column still has to be resolvable and still counts against the
    limits, but neither the header nor the cell may show up. Sorting by it has to
    keep working, otherwise "sort by something the reader should not see" -- the
    reason the flag exists -- silently stops working.
    """
    document = definition(
        base="order",
        columns=[
            {"field": "order.code"},
            {"field": "order.pending_sum", "hidden": True},
        ],
        sorting=[{"field": "order.pending_sum", "direction": "desc"}],
    )
    headers, rows = run_report(document, world.event)
    assert headers == ["Order code"]
    # 26.00, 23.00, 15.00, then the two orders that owe 0.00, then -7.00. PAID1
    # and CANC5 tie at 0.00 and are separated by the primary-key tiebreaker the
    # compiler always appends -- PAID1 was created first. Without that tiebreaker
    # this line would be flaky, which is exactly why it is asserted here.
    assert [row[0] for row in rows] == [
        "PART2",
        "PEND3",
        "EXPI4",
        "PAID1",
        "CANC5",
        "OVER6",
    ]
    assert all(len(row) == 1 for row in rows)


@pytest.mark.django_db
def test_a_variation_and_an_empty_category_render_as_themselves(world):
    """Two easy-to-miss cells: a product variation and a missing relation.

    ``merch`` has no category on purpose. An attribute renderer that does not
    handle a missing relation raises here; one that swallows every exception
    turns the *whole* column into blanks, which is the failure nobody notices.
    """
    document = definition(
        base="orderposition",
        columns=columns(
            "position.code", "item.internal_name", "item.category", "variation.value"
        ),
        sorting=[
            {"field": "order.code", "direction": "asc"},
            {"field": "position.positionid", "direction": "asc"},
        ],
    )
    _, rows = run_report(document, world.event)
    indexed = rows_by_code(rows)
    assert indexed["PAID1-1"][1:] == ["ticket", "Tickets", None]
    assert indexed["PAID1-2"][1:] == ["workshop", "Extras", "Beginner"]
    assert indexed["EXPI4-1"][1:] == ["merch", None, None]


@pytest.mark.django_db
def test_the_voucher_and_the_invoice_address_reach_their_columns(world):
    """One position has a voucher, one order has an invoice address.

    Both are nullable relations reached through different bases, and both are in
    ``SPEC.md`` F4's field list. A row with ``None`` in all four cells would pass
    a "does it run" test.
    """
    document = definition(
        base="orderposition",
        columns=columns(
            "position.code",
            "voucher.code",
            "voucher.tag",
            "invoice_address.company",
            "invoice_address.country",
        ),
        sorting=[
            {"field": "order.code", "direction": "asc"},
            {"field": "position.positionid", "direction": "asc"},
        ],
    )
    _, rows = run_report(document, world.event)
    indexed = rows_by_code(rows)
    assert indexed["PEND3-1"][1:3] == ["EARLYBIRD", "promo"]
    assert indexed["PAID1-1"][1:3] == [None, None]
    assert indexed["PAID1-1"][3] == "Analytical Engines Ltd"
    assert str(indexed["PAID1-1"][4]) == "GB"
    assert indexed["PEND3-1"][3] is None


# ===========================================================================
# 2b. The findings: two closed (T-001, T-002), one new (T-004)
# ===========================================================================
#
# All of them are cross-agent gaps: each component behaves as its own suite
# says, and the defect only exists at the seam. Recorded in handoff/blockers.md.
#
# T-001 and T-002 were ``xfail(strict=True)`` from wave 3 until 2026-08-03, when
# ``exporter-dev``/``frontend-dev`` and ``registry-dev``/``query-dev`` fixed
# them. The markers are gone and the assertions are unchanged: a reproducer that
# has been verified to fall over with the fix neutralised is the best regression
# guard the finding can leave behind, and rewriting its assertion at the moment
# it turns green would throw exactly that away. What each of them is *worth* is
# recorded in the docstring, so that a later failure is read as "the fix was
# undone" rather than as "some test about formats broke".
#
# The tests around them are new and adversarial: they exist because a fix that
# makes one reproducer pass is not the same thing as a fix that holds. See
# handoff/status/test-engineer.md, "Verifikation T-001 bis T-003".
#
# T-004 is new, found while verifying T-002, and is a strict xfail for the same
# reason T-002 was. To watch the open one fail for real:
#
#     pytest tests/test_integration.py -k finding --runxfail


@pytest.mark.django_db
def test_finding_a_column_format_chosen_in_the_editor_reaches_the_export(
    registered, user_with_perms, world
):
    """The same column, two date styles, two different files. **T-001, closed.**

    ``ColumnFormat`` is part of the frozen contract, the editor offers it per
    datatype and the live preview honoured it -- and until 2026-08-03 nothing in
    the export path did, so "date only" showed a date on screen and a full
    timestamp in the file. ``exporters.py`` now holds the single renderer
    (``format_cell_value``/``format_export_cell``), the exporter applies it via
    ``CustomReportExporter._cell_formats()`` and ``views/api.py`` imports the
    same function instead of keeping its own copy.

    The measured before/after, both lines from one and the same report::

        before:  iso -> "2026-04-24 09:00:00+00:00"
                 date_only -> "2026-04-24 09:00:00+00:00"
        after:   iso -> "2026-04-24T09:00:00+00:00"
                 date_only -> "2026-04-24"

    Verified to fail again at this assertion with the renderer neutralised at run
    time (``exporters.format_export_cell`` replaced by the identity, no
    production code touched): both styles collapse back onto the first pair of
    lines above.
    """
    rendered = {}
    for style in ("iso", "date_only"):
        document = definition(
            base="order",
            columns=[
                {"field": "order.code"},
                {"field": "order.datetime", "format": {"date_style": style}},
            ],
            filters={
                "op": "and",
                "children": [
                    {"field": "order.code", "operator": "exact", "value": "PAID1"}
                ],
            },
        )
        with scopes_disabled():
            report = ReportDefinition.objects.filter(
                event=world.event, identifier="fmt"
            ).first()
            if report is None:
                report = ReportDefinition.objects.create(
                    event=world.event,
                    name="Format",
                    identifier="fmt",
                    definition=document,
                )
            else:
                report.definition = document
                report.save()
        _, lines = export_csv(world.event, user_with_perms, "fmt")
        rendered[style] = lines[1]

    assert rendered["date_only"] != rendered["iso"], "both styles produced %r" % (
        rendered["iso"],
    )
    # Not just "different": the two strings a reader would expect. "Different"
    # alone would also be satisfied by a renderer that mangles both.
    assert rendered["iso"] == ["PAID1", "2026-04-24T09:00:00+00:00"]
    assert rendered["date_only"] == ["PAID1", "2026-04-24"]


@pytest.mark.django_db
def test_finding_an_aggregated_money_column_keeps_its_two_decimal_places(
    registered, user_with_perms, world
):
    """``order.total`` said ``23.50``, ``payment.sum_confirmed`` said ``20.5``.
    **T-002, closed.**

    Django's SQLite backend quantises a ``DecimalField`` to its ``decimal_places``
    only when the expression is a plain column; for a ``Subquery``/``Coalesce`` it
    hands the raw value through (``django/db/backends/sqlite3/operations.py``,
    ``get_decimalfield_converter``). PostgreSQL keeps the scale of
    ``numeric(13,2)`` through ``SUM``, so **the same report produced two
    different files on two installations** -- the same class of problem as the
    ``nulls_last`` divergence ``query-dev`` guarded against deliberately.

    Fixed on 2026-08-03 by ``registry/annotations.py::MoneyField`` (a
    ``DecimalField`` with a ``from_db_value`` that quantises, so the guarantee
    hangs on the output field rather than on a backend branch) and, for the
    aggregate the *user* picks, by ``query/relations.py::aggregate_expression``.

    Verified to fail again at the assertion below with either half of the fix
    neutralised at run time, and each half fails on its own columns -- which is
    what says the two fixes are not covering for each other::

        MoneyField.from_db_value removed  -> ['23.50', '20.5',  '23.50']
        aggregate_expression bypassed     -> ['23.50', '20.50', '23.5']

    The file is checked as well as the values: ``Decimal("23.5") ==
    Decimal("23.50")`` is ``True`` in Python, so the defect is only ever visible
    in the *characters* an export writes, and a test that compared ``Decimal``
    objects would have stayed green through all of it.
    """
    with scopes_disabled():
        order = factories.make_order(
            world.event, "CENTS", Order.STATUS_PENDING, Decimal("23.50")
        )
        factories.add_position(order, world.catalog.ticket, Decimal("23.50"), 1)
        factories.add_payment(order, Decimal("20.50"))

    document = definition(
        base="order",
        columns=columns(
            "order.code",
            "order.total",
            "payment.sum_confirmed",
            ("position.price", "sum"),
        ),
        filters={
            "op": "and",
            "children": [
                {"field": "order.code", "operator": "exact", "value": "CENTS"}
            ],
        },
    )
    _, rows = run_report(document, world.event)
    cells = [str(cell) for cell in rows[0][1:]]
    assert cells == ["23.50", "20.50", "23.50"], cells

    # And once through the real file, because that is where the finding lived.
    store_report(world.event, document, identifier="cents")
    _, lines = export_csv(world.event, user_with_perms, "cents")
    assert lines[1] == ["CENTS", "23.50", "20.50", "23.50"]


@pytest.mark.django_db
def test_every_money_column_the_editor_offers_keeps_two_decimal_places(world):
    """All fourteen money cells of one row, not just the three of the finding.

    A fix verified against its own reproducer proves the reproducer. This is the
    enumeration instead: every field the registry declares
    ``DataType.MONEY`` for, in every aggregate it allows, in a single row -- the
    four registry expressions (``payment.sum_confirmed``, ``refund.sum_done``,
    ``order.pending_sum``, ``position.net_price``) plus the twelve
    field-by-aggregate combinations the editor lets a user pick, plus
    ``order.total`` as the plain column that always kept its scale and is
    therefore the yardstick the others have to match *inside the same row*.

    The numbers, by pencil, for the order built below (23.50 + 10.00 gross,
    3.50 + 1.60 tax, 20.50 paid, 0.50 refunded, product list price 23.00):

    ======================  =======  ======================================
    column                  value    why
    ======================  =======  ======================================
    order.total             23.50    stored on the order
    payment.sum_confirmed   20.50    one confirmed payment
    refund.sum_done         0.50     one done refund
    order.pending_sum       3.50     23.50 - 20.50 + 0.50
    sum(position.price)     33.50    23.50 + 10.00
    min/max(position.price) 10.00 /  the two positions
                            23.50
    avg(position.price)     16.75    33.50 / 2
    sum(position.tax_value) 5.10     3.50 + 1.60
    avg(position.tax_value) 2.55     5.10 / 2
    sum(item.default_price) 46.00    the same product twice, 23.00 each
    ======================  =======  ======================================

    ``join`` is not in the table because no money field offers it, and
    ``count``/``count_distinct`` are not because a cardinality has no scale to
    lose -- both checked against the registry rather than assumed.
    """
    with scopes_disabled():
        order = factories.make_order(
            world.event, "MONEY", Order.STATUS_PENDING, Decimal("23.50")
        )
        factories.add_position(
            order, world.catalog.ticket, Decimal("23.50"), 1, tax_value=Decimal("3.50")
        )
        factories.add_position(
            order, world.catalog.ticket, Decimal("10.00"), 2, tax_value=Decimal("1.60")
        )
        factories.add_payment(order, Decimal("20.50"))
        factories.add_refund(order, Decimal("0.50"))

    document = definition(
        base="order",
        columns=columns(
            "order.code",
            "order.total",
            "payment.sum_confirmed",
            "refund.sum_done",
            "order.pending_sum",
            ("position.price", "sum"),
            ("position.price", "min"),
            ("position.price", "max"),
            ("position.price", "avg"),
            ("position.tax_value", "sum"),
            ("position.tax_value", "min"),
            ("position.tax_value", "max"),
            ("position.tax_value", "avg"),
            ("item.default_price", "sum"),
            ("item.default_price", "min"),
            ("item.default_price", "max"),
            ("item.default_price", "avg"),
        ),
        filters={
            "op": "and",
            "children": [
                {"field": "order.code", "operator": "exact", "value": "MONEY"}
            ],
        },
    )
    _, rows = run_report(document, world.event)
    assert [str(cell) for cell in rows[0]] == [
        "MONEY",
        "23.50",
        "20.50",
        "0.50",
        "3.50",
        "33.50",
        "10.00",
        "23.50",
        "16.75",
        "5.10",
        "1.60",
        "3.50",
        "2.55",
        "46.00",
        "23.00",
        "23.00",
        "23.00",
    ]


@pytest.mark.django_db
def test_the_average_of_a_money_column_is_rounded_to_cents(world):
    """43,00 over three positions is 14,33 -- a decision, not a rounding error.

    ``query-dev`` chose to quantise ``AVG`` on a money field, and that deserves a
    test of its own rather than hiding inside the table above, because it is the
    one place in T-002's fix where a *value* changes rather than its notation.
    Measured with the quantisation bypassed at run time, the same cell reads
    ``Decimal("14.3333333333333")`` -- thirteen digits of SQLite's float path,
    which PostgreSQL would answer differently. That is the argument: the
    unrounded number is an artefact of the installation, so it is not more
    precise, only less comparable. Anyone who needs the exact quotient has
    ``sum`` and ``count`` as two columns.
    """
    with scopes_disabled():
        order = factories.make_order(
            world.event, "THIRD", Order.STATUS_PENDING, Decimal("43.00")
        )
        for index, price in enumerate(("10.00", "13.00", "20.00"), start=1):
            factories.add_position(order, world.catalog.ticket, Decimal(price), index)
    document = definition(
        base="order",
        columns=columns(
            "order.code",
            ("position.price", "avg"),
            ("position.price", "sum"),
            ("position.price", "count"),
        ),
        filters={
            "op": "and",
            "children": [
                {"field": "order.code", "operator": "exact", "value": "THIRD"}
            ],
        },
    )
    _, rows = run_report(document, world.event)
    assert [str(cell) for cell in rows[0]] == ["THIRD", "14.33", "43.00", "3"]


@pytest.mark.django_db
def test_an_aggregate_over_no_rows_stays_empty_instead_of_becoming_zero(world):
    """The other half of the quantisation: ``None`` must survive it.

    A converter that runs on every value of a money expression is one ``if`` away
    from turning "this order has no positions" into ``0.00``, and an empty cell
    and a zero mean different things in an accounting export -- an order with no
    positions is not an order worth nothing. ``count`` is the deliberate
    exception and stays ``0``, because a missing count really is zero
    (``_COALESCE_TO_ZERO`` in ``query/relations.py``).
    """
    with scopes_disabled():
        factories.make_order(
            world.event, "NOPOS", Order.STATUS_PENDING, Decimal("0.00")
        )
    document = definition(
        base="order",
        columns=columns(
            "order.code",
            ("position.price", "sum"),
            ("position.price", "min"),
            ("position.price", "max"),
            ("position.price", "avg"),
            ("position.price", "count"),
        ),
        filters={
            "op": "and",
            "children": [
                {"field": "order.code", "operator": "exact", "value": "NOPOS"}
            ],
        },
    )
    _, rows = run_report(document, world.event)
    assert rows[0] == ["NOPOS", None, None, None, None, 0]


@pytest.mark.django_db
@pytest.mark.xfail(
    strict=True,
    reason="T-004: aggregate_expression pins the scale for DataType.MONEY only, "
    "so an aggregate over position.tax_rate (DataType.DECIMAL, decimal(7,2) in "
    "pretix) writes '19' where the plain column writes '19.00'",
)
def test_finding_an_aggregated_decimal_column_keeps_its_scale(world):
    """T-002 one datatype over: ``position.tax_rate``, aggregated.

    Found while enumerating the money paths for the T-002 verification. The fix
    for T-002 keys on ``DataType.MONEY``
    (``query/relations.py::aggregate_expression``), and ``DataType.DECIMAL`` goes
    through the same ``Sum``/``Min``/``Max``/``Avg`` with the model's plain
    ``DecimalField`` as its output field -- which is exactly the state money was
    in before. ``position.tax_rate`` is a core registry field, the editor offers
    it with all six aggregates, and pretix declares it
    ``DecimalField(max_digits=7, decimal_places=2)``
    (``pretix/base/models/orders.py:2558``).

    Measured for one order with a 19,00 % and a 7,00 % position::

        base orderposition, plain column   -> "19.00", "7.00"
        base order, min/max/sum/avg        -> "7", "19", "26", "13"

    Same two symptoms as T-002, and both of them: one file that disagrees with
    itself (``Tax rate`` says ``19.00``, ``Highest tax rate`` says ``19``), and
    two installations that disagree with each other, because PostgreSQL keeps the
    scale of ``numeric(7,2)`` through ``SUM`` and SQLite does not.

    Deliberately *not* the same one-line fix: ``MoneyField`` may hard-code two
    decimal places because every money column in pretix has two, while
    ``DataType.DECIMAL`` covers fields of different scales, and the registry does
    not declare a scale today. Whoever picks this up decides between "carry the
    scale in ``ReportField``" and "quantise ``DECIMAL`` to the model field's own
    ``decimal_places``". Owner: ``query-dev`` with ``registry-dev``; severity
    lower than T-002, because a tax rate is not an amount anybody adds up.
    """
    with scopes_disabled():
        order = factories.make_order(
            world.event, "TAXRT", Order.STATUS_PENDING, Decimal("20.00")
        )
        factories.add_position(
            order, world.catalog.ticket, Decimal("10.00"), 1, tax_rate=Decimal("19.00")
        )
        factories.add_position(
            order, world.catalog.ticket, Decimal("10.00"), 2, tax_rate=Decimal("7.00")
        )
    document = definition(
        base="order",
        columns=columns(
            "order.code",
            ("position.tax_rate", "min"),
            ("position.tax_rate", "max"),
            ("position.tax_rate", "sum"),
        ),
        filters={
            "op": "and",
            "children": [
                {"field": "order.code", "operator": "exact", "value": "TAXRT"}
            ],
        },
    )
    _, rows = run_report(document, world.event)
    assert [str(cell) for cell in rows[0]] == ["TAXRT", "7.00", "19.00", "26.00"]


# ===========================================================================
# 2c. What the T-001 fix has to keep promising
# ===========================================================================
#
# Written while verifying the fix, not while reporting the finding. Each of them
# is a place where a renderer that satisfies the reproducer could still be
# wrong: the wrong column, the wrong output format, the wrong event, or a
# preview that has quietly drifted apart from the file again.


#: Every style the contract offers, paired with a column of the matching
#: datatype. Written out rather than derived from the enums, so that adding a
#: style to the contract makes this list visibly incomplete instead of silently
#: covering it.
STYLE_MATRIX = (
    ("order.datetime", {"date_style": "iso"}, "2026-04-24T09:00:00+00:00"),
    ("order.datetime", {"date_style": "date_only"}, "2026-04-24"),
    ("order.datetime", {"date_style": "time_only"}, "09:00"),
    ("order.datetime", {"date_style": "short"}, "2026-04-24 09:00"),
    ("order.datetime", {"date_style": "long"}, "Friday, 24 April 2026 09:00"),
    ("order.total", {"number_style": "raw"}, "33.00"),
    ("order.total", {"number_style": "localized"}, "33.00"),
    ("order.total", {"number_style": "currency"}, "€33.00"),
    ("payment.sum_confirmed", {"number_style": "currency"}, "€33.00"),
    ("order.testmode", {"boolean_style": "yes_no"}, "No"),
    ("order.testmode", {"boolean_style": "true_false"}, "false"),
    ("order.testmode", {"boolean_style": "one_zero"}, "0"),
)


@pytest.mark.django_db
def test_every_column_format_renders_the_same_string_in_the_preview_and_in_the_file(
    wired_urls, registered, client_with_perms, user_with_perms, world
):
    """Twelve styles, one row, two paths, the same twelve strings.

    The reproducer for T-001 uses two date styles; a renderer wired into the
    export for those two and not for the rest would satisfy it. So this walks the
    whole contract: five ``DateStyle``, three ``NumberStyle``, three
    ``BooleanStyle``, once through ``api/preview/`` over HTTP and once through
    ``ListExporter`` into a CSV, compared cell by cell **and** against the
    literal strings above.

    Both halves are load-bearing. Comparing preview to export alone would be
    satisfied by two renderers that are equally broken -- which is precisely the
    shape T-001 had, only inverted. Comparing to the literals alone would not
    notice the preview drifting off again, which is what
    ``get_cell_renderer()`` exists to prevent.
    """
    cells = [{"field": "order.code"}] + [
        {"field": key, "format": fmt} for key, fmt, _expected in STYLE_MATRIX
    ]
    document = definition(
        base="order",
        columns=cells,
        filters={
            "op": "and",
            "children": [
                {"field": "order.code", "operator": "exact", "value": "PAID1"}
            ],
        },
    )
    store_report(world.event, document, identifier="styles")

    response = post_json(
        client_with_perms,
        url_for("api.preview", world.event),
        {"definition": document},
    )
    assert response.status_code == 200
    preview = response.json()["rows"][0]

    _, lines = export_csv(world.event, user_with_perms, "styles")
    expected = ["PAID1"] + [text for _key, _fmt, text in STYLE_MATRIX]
    assert lines[1] == expected
    assert preview == expected


@pytest.mark.django_db
def test_a_hidden_column_does_not_shift_the_formats_of_the_columns_behind_it(
    registered, user_with_perms, world
):
    """A format is paired with a column by position -- among the *visible* ones.

    ``CompiledReport.columns`` has hidden columns dropped already, so pairing the
    exporter's rows against ``definition.columns`` index for index would apply
    every format one column too far to the left as soon as a report hides
    anything. That is a silent wrong answer, not a crash, and hidden columns are
    common: they are how a report filters or sorts by something it does not
    print.

    The hidden column here sits *between* the two visible ones and carries a
    format of its own, so an off-by-one would be visible twice over -- the date
    column would come out unformatted and the code column would be handed a
    date style.
    """
    document = definition(
        base="order",
        columns=[
            {"field": "order.code"},
            {
                "field": "order.total",
                "hidden": True,
                "format": {"number_style": "currency"},
            },
            {"field": "order.datetime", "format": {"date_style": "date_only"}},
        ],
        sorting=[{"field": "order.total", "direction": "desc"}],
        filters={
            "op": "and",
            "children": [
                {"field": "order.code", "operator": "exact", "value": "PAID1"}
            ],
        },
    )
    store_report(world.event, document, identifier="hidden-fmt")
    _, lines = export_csv(world.event, user_with_perms, "hidden-fmt")
    assert lines == [
        ["Order code", "Order date"],
        ["PAID1", "2026-04-24"],
    ]


@pytest.mark.django_db
def test_a_multi_event_export_formats_each_events_rows_in_that_events_timezone(
    registered, user_with_perms, organizer
):
    """Two events, two formats, two time zones, one file.

    The organizer-level export compiles once per event and then writes all the
    rows into a single table. Both inputs to the renderer therefore have to be
    re-read per event and not hoisted out of the loop: the ``ColumnFormat`` comes
    from *that* event's report (identifiers are unique per event, so the same
    identifier can hold two different definitions), and the timezone an aware
    datetime is localised into comes from *that* event.

    Same instant in both rows -- 2026-04-24 09:00 UTC. Berlin is +02:00 in April
    and Auckland +12:00, so a renderer using the server zone, the first event's
    zone or the first event's format would produce a table in which one of these
    two lines is wrong, and nothing about the file would say so.
    """
    for slug, timezone, style in (
        ("berlin", "Europe/Berlin", "short"),
        ("auckland", "Pacific/Auckland", "iso"),
    ):
        event = factories.make_event(
            organizer, slug=slug, name=slug.title(), timezone=timezone
        )
        with scopes_disabled():
            catalog = factories.make_catalog(event)
            order = factories.make_order(
                event, slug[:5].upper(), Order.STATUS_PAID, Decimal("10.00")
            )
            factories.add_position(order, catalog.ticket, Decimal("10.00"), 1)
        store_report(
            event,
            definition(
                base="order",
                columns=[
                    {"field": "order.code"},
                    {"field": "order.datetime", "format": {"date_style": style}},
                ],
            ),
            identifier="zones",
        )

    from pretix.base.services.export import init_organizer_exporters

    with scope(organizer=organizer):
        exporter = next(
            ex
            for ex in init_organizer_exporters(
                organizer=organizer, user=user_with_perms
            )
            if ex.identifier == exporters.CustomReportExporter.identifier
        )
        rows = export_rows(
            exporter,
            {"_format": "default", contracts.EXPORT_FORM_REPORT_KEY: "zones"},
        )

    assert rows == [
        ["Event slug", "Event name", "Order code", "Order date"],
        ["auckland", "Auckland", "AUCKL", "2026-04-24T21:00:00+12:00"],
        ["berlin", "Berlin", "BERLI", "2026-04-24 11:00"],
    ]


@pytest.mark.django_db
def test_a_scheduled_export_mails_a_file_with_the_chosen_column_formats(
    registered, user_with_perms, world
):
    """The unattended path, because that is where nobody looks at the screen.

    A column format is set in the editor and read back out of the **mail
    attachment** a ``ScheduledEventExport`` produces, through the real
    ``run_scheduled_exports`` receiver. The whole point of T-001 was that a user
    trusts what the preview showed; for a scheduled report there is no preview at
    the moment it runs, only the file that arrives, and it is the one path where
    a formatting bug can survive for months without anybody noticing.
    """
    from django.core import mail as djmail

    store_report(
        world.event,
        definition(
            base="order",
            columns=[
                {"field": "order.code"},
                {"field": "order.datetime", "format": {"date_style": "date_only"}},
                {"field": "order.total", "format": {"number_style": "currency"}},
            ],
            sorting=[{"field": "order.code", "direction": "asc"}],
        ),
        identifier="nightly",
    )
    with scopes_disabled():
        ScheduledEventExport.objects.create(
            event=world.event,
            owner=user_with_perms,
            export_identifier=exporters.CustomReportExporter.identifier,
            export_form_data={
                "_format": "default",
                contracts.EXPORT_FORM_REPORT_KEY: "nightly",
            },
            locale="en",
            mail_additional_recipients="",
            mail_subject="Nightly",
            mail_template="Attached.",
            schedule_rrule="DTSTART:20260101T000000\nRRULE:FREQ=DAILY",
            schedule_rrule_time="04:00:00",
            schedule_next_run="2026-01-01T04:00:00Z",
            error_counter=0,
        )

    djmail.outbox = []
    run_scheduled_exports(None)

    assert len(djmail.outbox) == 1
    attachments = djmail.outbox[0].attachments
    assert len(attachments) == 1
    name, content, _mime = attachments[0]
    assert name.endswith(".csv")
    if isinstance(content, bytes):  # pragma: no cover - depends on the backend
        content = content.decode("utf-8-sig")
    assert list(csv.reader(io.StringIO(content.lstrip("﻿")))) == [
        ["Order code", "Order date", "Order total"],
        ["CANC5", "2026-04-26", "€23.00"],
        ["EXPI4", "2026-05-02", "€15.00"],
        ["OVER6", "2026-05-03", "€23.00"],
        ["PAID1", "2026-04-24", "€33.00"],
        ["PART2", "2026-04-29", "€46.00"],
        ["PEND3", "2026-05-01", "€23.00"],
    ]


@pytest.mark.django_db
def test_a_date_column_reaches_the_xlsx_path_with_and_without_a_style(
    registered, user_with_perms, organizer
):
    """XLSX is the output format that can reject a value the CSV accepts.

    Two things at once, and they belong together because one is the other's
    counter-check:

    * A **styled** datetime arrives as the string the style asks for -- the same
      claim as for CSV, on the path where the exporter also has to decide *not*
      to touch anything else.
    * An **unstyled** datetime arrives as a real ``datetime`` a spreadsheet can
      compute with, and does so at all: openpyxl refuses a timezone-aware value
      outright ("Excel does not support timezones in datetimes"), and
      ``ListExporter._render_xlsx`` passes our cells to ``ws.append`` unchanged.
      Before ``as_spreadsheet_value()`` this raised ``TypeError`` inside a Celery
      task -- not an ``ExportError`` naming the column, but five retries and
      "Internal Error", after which the schedule drops out of the periodic query.
      Found by ``exporter-dev`` alongside T-001 and pinned here at the level
      where it hurt: a whole export, not a function.

    The instant is 09:00 UTC and the event is in Auckland (+12:00 in April), so
    the naive value written into the sheet has to read 21:00. A spreadsheet cell
    cannot say which zone it is in, so writing UTC would be wrong by twelve hours
    and look perfectly plausible. The money cell is checked in the same row for
    the opposite reason: an unstyled number must stay a **number**, not become a
    string that only looks like one, which is the whole reason the formatting
    lives in the exporter and not in the compiler.
    """
    event = factories.make_event(
        organizer, slug="sheet", name="Sheet", timezone="Pacific/Auckland"
    )
    with scopes_disabled():
        catalog = factories.make_catalog(event)
        order = factories.make_order(
            event, "XLS01", Order.STATUS_PAID, Decimal("23.50")
        )
        factories.add_position(order, catalog.ticket, Decimal("23.50"), 1)
    store_report(
        event,
        definition(
            base="order",
            columns=[
                {"field": "order.code"},
                {"field": "order.datetime"},
                {"field": "order.datetime", "format": {"date_style": "date_only"}},
                {"field": "order.total"},
                {"field": "order.total", "format": {"number_style": "currency"}},
            ],
        ),
        identifier="sheet",
    )

    rows = export_xlsx(event, user_with_perms, "sheet")
    assert rows[0] == [
        "Order code",
        "Order date",
        "Order date",
        "Order total",
        "Order total",
    ]
    code, unstyled_date, styled_date, unstyled_total, styled_total = rows[1]
    assert code == "XLS01"
    assert unstyled_date == dt.datetime(2026, 4, 24, 21, 0)
    assert unstyled_date.tzinfo is None
    assert styled_date == "2026-04-24"
    assert unstyled_total == 23.5 and not isinstance(unstyled_total, str)
    assert styled_total == "€23.50"


@pytest.mark.django_db
@pytest.mark.xfail(
    strict=True,
    reason="T-005: format_cell_value and as_spreadsheet_value resolve "
    "event.timezone per cell, so a formatted date column costs one settings "
    "lookup per row where one per export would do",
)
def test_finding_the_export_resolves_the_event_timezone_once_not_once_per_row(
    monkeypatch, registered, user_with_perms, world
):
    """A formatted date column reads ``Event.timezone`` once per row.

    Found while measuring what the T-001 fix costs, and it is the only thing in
    that fix that scales with the row count. ``Event.timezone`` is not an
    attribute: it is ``pytz_deprecation_shim.timezone(self.settings.timezone)``
    (``pretix/base/models/event.py:233-235``), i.e. a hierarkey settings lookup
    that walks event -> organizer -> global defaults. ``_format_temporal()`` and
    ``as_spreadsheet_value()`` each call it for **every cell they touch**.

    Counted here rather than timed, because a count is deterministic and a wall
    clock on a shared machine is not. The baseline is whatever the compiler and
    the exporter need anyway; the question is only whether adding a *style* adds
    a fixed amount or a per-row amount, and the answer today is per-row:

    ======================  =========  ===========================
    report                  1 row      6 rows
    ======================  =========  ===========================
    unstyled date column    22 reads   22 reads   (flat, not ours)
    ``date_only`` on it     23 reads   28 reads   (+1 per row)
    ======================  =========  ===========================

    What it costs: measured on the load fixture, a 94.666-row / 22-column CSV
    export goes from 11,6 s to 50,4 s (x4,4) when three columns carry a style,
    and about 17 s of the XLSX export's 69,7 s is the same lookup inside
    ``as_spreadsheet_value``. One resolution is 178-345 us here against 1,8 us
    for the conversion it guards -- two orders of magnitude of pure overhead
    (``docs/performance.md`` 3.8). The absolute figure depends on the cache
    backend (pretix' test settings use ``DummyCache``); the *shape* does not.

    Not a correctness problem, and deliberately not a blocker: every value is
    right, and a scheduled export that takes fifty seconds instead of twelve
    still arrives. It is a fix of a handful of lines in somebody else's file, so
    it is a finding: resolve the zone once per event -- next to
    ``_cell_formats()``, which is already computed once per event for exactly
    this reason -- and hand it down. Owner: ``exporter-dev``.

    The assertion is deliberately loose (``<= 2``): the point is that the extra
    work must not grow with the report, not that it must be exactly zero.
    """
    from pretix.base.models import Event

    reads = []
    original = Event.timezone.fget
    monkeypatch.setattr(
        Event,
        "timezone",
        property(lambda self: (reads.append(self.pk), original(self))[1]),
    )

    def count_reads(identifier, fmt):
        store_report(
            world.event,
            definition(
                base="order",
                columns=[
                    {"field": "order.code"},
                    {"field": "order.datetime", "format": fmt},
                ],
            ),
            identifier=identifier,
        )
        reads.clear()
        _, lines = export_csv(world.event, user_with_perms, identifier)
        return len(lines) - 1, len(reads)

    plain_rows, plain_reads = count_reads("tz-plain", {})
    styled_rows, styled_reads = count_reads("tz-styled", {"date_style": "date_only"})

    assert plain_rows == styled_rows == 6
    assert styled_reads - plain_reads <= 2, (
        f"formatting {styled_rows} rows cost {styled_reads - plain_reads} extra "
        f"timezone resolutions ({plain_reads} -> {styled_reads})"
    )


# ===========================================================================
# 3. Edge cases
# ===========================================================================


@pytest.mark.django_db
def test_an_event_without_any_orders_exports_a_header_and_nothing_else(
    registered, user_with_perms, organizer
):
    """The empty event. A header row, zero data rows, no exception.

    This is the state a freshly copied event is in, and a scheduled export
    pointing at it must not fail -- an empty report is a legitimate answer.
    """
    empty = factories.make_event(organizer, slug="empty", name="Empty")
    with scopes_disabled():
        factories.make_catalog(empty)
        ReportDefinition.objects.create(
            event=empty,
            name="Codes",
            identifier="codes",
            definition=definition(
                base="order", columns=columns("order.code", "order.total")
            ),
        )
    _, lines = export_csv(empty, user_with_perms, "codes")
    assert lines == [["Order code", "Order total"]]


@pytest.mark.django_db
def test_an_order_without_positions(world, organizer):
    """A row on base ``order``, no row on base ``orderposition``.

    An order without positions exists in practice (every position canceled, or a
    fee-only order). On base ``order`` it must still produce a row -- dropping it
    would make the report disagree with the order list -- with ``0`` for counts
    and ``None`` for a ``SUM`` over nothing. On base ``orderposition`` there is
    nothing to show, which is not the same thing as an error.
    """
    with scopes_disabled():
        factories.make_order(
            world.event, "NOPOS", Order.STATUS_PENDING, Decimal("0.00")
        )

    document = definition(
        base="order",
        columns=columns(
            "order.code",
            "order.position_count",
            ("position.price", "sum"),
            ("position.positionid", "count"),
        ),
    )
    _, rows = run_report(document, world.event)
    assert rows_by_code(rows)["NOPOS"] == ["NOPOS", 0, None, 0]

    document = definition(
        base="orderposition", columns=columns("order.code", "position.positionid")
    )
    _, rows = run_report(document, world.event)
    assert "NOPOS" not in {row[0] for row in rows}


@pytest.mark.django_db
def test_a_deleted_question_leaves_the_report_openable_and_names_the_key(
    wired_urls, registered, client_with_perms, user_with_perms, world
):
    """Delete a question that a saved report uses. Three things must happen.

    1. the stored row survives -- an unresolvable key is a legal saved state
       (``models.py``), otherwise the report could never be repaired,
    2. the editor still opens it and *says* which key is broken, at the right
       document path,
    3. running it fails with a message that names the key, rather than an
       ``Internal Error`` mail and five Celery retries.

    Deleting a question is the ordinary case, not an exotic one: identifiers are
    user-editable and a renamed question is indistinguishable from a deleted one.
    """
    with scopes_disabled():
        report = ReportDefinition.objects.create(
            event=world.event,
            name="Sizes",
            identifier="sizes",
            definition=definition(
                base="orderposition",
                columns=columns("position.code", "answer.tshirt-size"),
            ),
        )
        world.questions["tshirt-size"].delete()
    registry_cache.clear_local_cache()

    with scopes_disabled():
        assert ReportDefinition.objects.filter(pk=report.pk).exists()

    response = client_with_perms.get(
        url_for("editor.edit", world.event, identifier=report.identifier)
    )
    assert response.status_code == 200

    response = post_json(
        client_with_perms,
        url_for("api.validate", world.event),
        {"definition": report.definition},
    )
    warnings_ = response.json()["warnings"]
    assert [warning["path"] for warning in warnings_] == ["columns[1]"]
    assert "answer.tshirt-size" in json.dumps(warnings_)

    with scope(organizer=world.event.organizer):
        exporter = init_event_exporter(
            identifier=exporters.CustomReportExporter.identifier,
            event=world.event,
            user=user_with_perms,
        )
        with pytest.raises(ExportError) as excinfo:
            export_rows(
                exporter,
                {
                    "_format": "default",
                    contracts.EXPORT_FORM_REPORT_KEY: report.identifier,
                },
            )
    assert "answer.tshirt-size" in str(excinfo.value)


@pytest.mark.django_db
def test_a_subevent_column_on_an_event_without_a_series(world, organizer):
    """``subevent.*`` on a non-series event: empty cells, not an error.

    The field exists on every event -- ``OrderPosition.subevent`` is nullable and
    the registry does not make the column conditional -- so the honest result is
    ``None`` everywhere. The counter-example is in the same test: a real series
    fills the same column, which proves the ``None`` above comes from the data
    and not from a column that never worked.
    """
    document = definition(
        base="orderposition",
        columns=columns("position.code", "subevent.name", "subevent.date_from"),
    )
    _, rows = run_report(document, world.event)
    assert rows, "the reference event does have positions"
    assert all(row[1] is None and row[2] is None for row in rows)

    series = factories.make_event(
        organizer, slug="series", name="Series", has_subevents=True
    )
    with scopes_disabled():
        catalog = factories.make_catalog(series)
        dates = factories.make_subevents(series, 2)
        order = factories.make_order(
            series, "SUB01", Order.STATUS_PAID, Decimal("46.00")
        )
        factories.add_position(
            order, catalog.ticket, Decimal("23.00"), 1, subevent=dates[0]
        )
        factories.add_position(
            order, catalog.ticket, Decimal("23.00"), 2, subevent=dates[1]
        )
    registry_cache.clear_local_cache()

    document = definition(
        base="orderposition",
        columns=columns("position.code", "subevent.name", "subevent.location"),
        sorting=[
            {"field": "order.code", "direction": "asc"},
            {"field": "position.positionid", "direction": "asc"},
        ],
    )
    _, rows = run_report(document, series)
    assert rows == [
        ["SUB01-1", "Date 1", "Room 1"],
        ["SUB01-2", "Date 2", "Room 2"],
    ]


@pytest.mark.django_db
def test_a_canceled_order_is_a_row_that_owes_nothing(world):
    """CANC5 is a row, has no live positions and an outstanding amount of 0.00.

    Three separate claims, and the third one is the one a naive implementation
    gets wrong: ``total - paid + refunded`` is ``23 - 23 + 23 = 23`` unless the
    canceled special case from ``Order.pending_sum`` is honoured, which replaces
    the total with zero.
    """
    document = definition(
        base="order",
        columns=columns(
            "order.code",
            "order.status",
            "order.cancellation_date",
            "order.pending_sum",
            "order.position_count",
        ),
        filters={
            "op": "and",
            "children": [{"field": "order.status", "operator": "exact", "value": "c"}],
        },
    )
    _, rows = run_report(document, world.event)
    assert len(rows) == 1
    row = rows[0]
    assert row[0] == "CANC5"
    assert row[1] == "c"
    assert row[2] is not None
    assert row[3] == Decimal("0.00")
    assert row[4] == 0


@pytest.mark.django_db
def test_an_order_with_two_hundred_positions(registered, user_with_perms, world):
    """200 positions on one order: 200 rows there, one row here, correct sums.

    ``MAX_COLUMNS`` is 200, so 200 positions is deliberately the same order of
    magnitude -- big enough that an accidental per-row query or a Python-side
    ``join`` over the wrong relation shows up, small enough for the normal test
    run. The sum is the arithmetic series 1.00 + 2.00 + ... + 200.00 = 20100.00.
    """
    with scopes_disabled():
        big = factories.make_order(
            world.event, "BIG01", Order.STATUS_PAID, Decimal("20100.00")
        )
        for index in range(1, 201):
            factories.add_position(big, world.catalog.ticket, Decimal(index), index)

    document = definition(
        base="order",
        columns=columns(
            "order.code",
            ("position.price", "sum"),
            ("position.positionid", "count"),
            "order.position_count",
        ),
        filters={
            "op": "and",
            "children": [
                {"field": "order.code", "operator": "exact", "value": "BIG01"}
            ],
        },
    )
    _, rows = run_report(document, world.event)
    assert rows == [["BIG01", Decimal("20100.00"), 200, 200]]

    document = definition(
        base="orderposition",
        columns=columns("position.code", "position.price"),
        filters={
            "op": "and",
            "children": [
                {"field": "order.code", "operator": "exact", "value": "BIG01"}
            ],
        },
    )
    _, rows = run_report(document, world.event)
    assert len(rows) == 200
    assert sum(row[1] for row in rows) == Decimal("20100.00")


@pytest.mark.django_db
def test_a_report_whose_row_limit_caps_the_result(world):
    """``options.row_limit`` caps rows **and** the reported count.

    Twenty rows next to a total of six hundred is a lie; the count query applies
    the same cap.
    """
    document = definition(
        base="orderposition",
        columns=columns("position.code"),
        options={"row_limit": 3},
    )
    parsed = contracts.validate_definition(document)
    compiler = ReportQueryCompiler(field_registry())
    with scopes_disabled():
        report = compiler.compile(parsed, world.event)
        assert len(list(report.iter_rows())) == 3
        assert report.count() == 3


@pytest.mark.django_db
def test_a_report_of_a_foreign_event_is_not_reachable_through_this_one(
    wired_urls, client_with_perms, world, second_event
):
    """The scoping claim, from the outside.

    Every lookup in the editor, the CRUD views and the exporter goes through
    ``event.custom_reports``. This asserts the consequence rather than the code:
    a report of event A is a 404 under event B's URLs, and B's exporter does not
    offer it.
    """
    with scopes_disabled():
        report = ReportDefinition.objects.create(
            event=world.event,
            name="Codes",
            identifier="codes",
            definition=definition(base="order", columns=columns("order.code")),
        )
    response = client_with_perms.get(
        url_for("event.reports.export", second_event, report=report.pk)
    )
    assert response.status_code == 404
    response = client_with_perms.get(
        url_for("editor.edit", second_event, identifier="codes")
    )
    assert response.status_code == 404


# ===========================================================================
# 4. Time: relative filters across time zones and DST
# ===========================================================================
#
# ``exporter-dev`` already froze the clock for the simple case. These go past
# the boundaries that the simple case cannot see: an event whose calendar day is
# a different day from the server's, and the two days a year that are not 24
# hours long.


def berlin_event(organizer, slug="berlin"):
    return factories.make_event(
        organizer,
        slug=slug,
        name="Berlin",
        timezone="Europe/Berlin",
        date_from=dt.datetime(2026, 6, 3, 9, 0, tzinfo=dt.timezone.utc),
    )


def order_at(event, code, moment, catalog=None):
    """An order placed at a precise instant, with one position."""
    with scopes_disabled():
        order = factories.make_order(
            event, code, Order.STATUS_PAID, Decimal("23.00"), placed=moment
        )
        if catalog is not None:
            factories.add_position(order, catalog.ticket, Decimal("23.00"), 1)
        return order


def codes_placed_today(event, now):
    document = definition(
        base="order",
        columns=columns("order.code"),
        filters={
            "op": "and",
            "children": [{"field": "order.datetime", "operator": "relative_today"}],
        },
    )
    _, rows = run_report(document, event, now=now)
    return sorted(row[0] for row in rows)


@pytest.mark.django_db
def test_relative_today_follows_the_event_timezone_across_the_date_line(organizer):
    """An Auckland event, a UTC server, and an instant where they disagree.

    Frozen instant: 2026-06-30 13:00 UTC. In Auckland (UTC+12 in June) that is
    2026-07-01 01:00, so "today" is the **first of July** for this organizer
    while the server still says the thirtieth of June.

    Three orders, by hand:

    * ``BEFORE`` 2026-06-30 11:00 UTC = 2026-06-30 23:00 NZST -> yesterday,
    * ``EARLY``  2026-06-30 12:30 UTC = 2026-07-01 00:30 NZST -> today,
    * ``LATE``   2026-07-01 10:00 UTC = 2026-07-01 22:00 NZST -> today.

    A report resolved in the server timezone returns ``BEFORE`` and ``EARLY``.
    That is the bug this test exists for, and it is invisible on a server whose
    timezone happens to match the organizer's.
    """
    event = factories.make_event(
        organizer,
        slug="auckland",
        name="Auckland",
        timezone="Pacific/Auckland",
        date_from=dt.datetime(2026, 7, 5, 9, 0, tzinfo=dt.timezone.utc),
    )
    order_at(event, "BEFOR", dt.datetime(2026, 6, 30, 11, 0, tzinfo=dt.timezone.utc))
    order_at(event, "EARLY", dt.datetime(2026, 6, 30, 12, 30, tzinfo=dt.timezone.utc))
    order_at(event, "LATER", dt.datetime(2026, 7, 1, 10, 0, tzinfo=dt.timezone.utc))

    reference = dt.datetime(2026, 6, 30, 13, 0, tzinfo=dt.timezone.utc)
    assert codes_placed_today(event, reference) == ["EARLY", "LATER"]


@pytest.mark.django_db
def test_the_current_month_can_be_a_different_month_in_the_event_timezone(organizer):
    """Same instant, two calendars: June on the server, July for the organizer.

    ``relative_current_month`` is the operator where a timezone mistake is worst,
    because it is off by up to a whole month rather than a day, and a monthly
    invoice report is exactly what people schedule.
    """
    event = factories.make_event(
        organizer,
        slug="auckland2",
        name="Auckland",
        timezone="Pacific/Auckland",
        date_from=dt.datetime(2026, 7, 5, 9, 0, tzinfo=dt.timezone.utc),
    )
    # 2026-06-30 12:30 UTC == 2026-07-01 00:30 NZST -> inside July.
    order_at(event, "JULY1", dt.datetime(2026, 6, 30, 12, 30, tzinfo=dt.timezone.utc))
    # 2026-06-30 11:00 UTC == 2026-06-30 23:00 NZST -> still June.
    order_at(event, "JUNE1", dt.datetime(2026, 6, 30, 11, 0, tzinfo=dt.timezone.utc))
    # 2026-07-31 11:30 UTC == 2026-07-31 23:30 NZST -> last minutes of July.
    order_at(event, "JULY2", dt.datetime(2026, 7, 31, 11, 30, tzinfo=dt.timezone.utc))
    # 2026-07-31 12:30 UTC == 2026-08-01 00:30 NZST -> already August.
    order_at(event, "AUGU1", dt.datetime(2026, 7, 31, 12, 30, tzinfo=dt.timezone.utc))

    document = definition(
        base="order",
        columns=columns("order.code"),
        filters={
            "op": "and",
            "children": [
                {"field": "order.datetime", "operator": "relative_current_month"}
            ],
        },
    )
    reference = dt.datetime(2026, 7, 15, 0, 0, tzinfo=dt.timezone.utc)
    _, rows = run_report(document, event, now=reference)
    assert sorted(row[0] for row in rows) == ["JULY1", "JULY2"]


@pytest.mark.django_db
def test_the_day_of_the_spring_dst_change_is_twenty_three_hours_long(organizer):
    """Europe/Berlin, 2026-03-29: 02:00 CET becomes 03:00 CEST.

    "Today" that day runs from 00:00+01:00 to 00:00+02:00 the next day -- 23
    hours, not 24. Orders placed in the first and the last half hour of that
    local day must be in, and the first half hour of the next local day must be
    out. An implementation that adds ``timedelta(days=1)`` to a UTC instant puts
    the 23:30 order in the wrong day.
    """
    event = berlin_event(organizer, slug="spring")
    # 2026-03-29 00:30+01:00 == 2026-03-28 23:30 UTC
    order_at(event, "FIRST", dt.datetime(2026, 3, 28, 23, 30, tzinfo=dt.timezone.utc))
    # 2026-03-29 23:30+02:00 == 2026-03-29 21:30 UTC
    order_at(event, "LASTX", dt.datetime(2026, 3, 29, 21, 30, tzinfo=dt.timezone.utc))
    # 2026-03-30 00:30+02:00 == 2026-03-29 22:30 UTC -> the next day
    order_at(event, "AFTER", dt.datetime(2026, 3, 29, 22, 30, tzinfo=dt.timezone.utc))
    # 2026-03-28 23:30+01:00 == 2026-03-28 22:30 UTC -> the day before
    order_at(event, "BEFOR", dt.datetime(2026, 3, 28, 22, 30, tzinfo=dt.timezone.utc))

    reference = dt.datetime(2026, 3, 29, 12, 0, tzinfo=dt.timezone.utc)
    assert codes_placed_today(event, reference) == ["FIRST", "LASTX"]


@pytest.mark.django_db
def test_the_day_of_the_autumn_dst_change_is_twenty_five_hours_long(organizer):
    """Europe/Berlin, 2026-10-25: 03:00 CEST becomes 02:00 CET.

    02:30 happens twice that day, once at +02:00 and once at +01:00. Both
    belong to the 25th, and a window built from a fixed 24-hour offset drops the
    second one.
    """
    event = berlin_event(organizer, slug="autumn")
    # 2026-10-25 02:30+02:00 == 2026-10-25 00:30 UTC (first pass)
    order_at(event, "PASS1", dt.datetime(2026, 10, 25, 0, 30, tzinfo=dt.timezone.utc))
    # 2026-10-25 02:30+01:00 == 2026-10-25 01:30 UTC (second pass)
    order_at(event, "PASS2", dt.datetime(2026, 10, 25, 1, 30, tzinfo=dt.timezone.utc))
    # 2026-10-25 23:30+01:00 == 2026-10-25 22:30 UTC
    order_at(event, "LASTX", dt.datetime(2026, 10, 25, 22, 30, tzinfo=dt.timezone.utc))
    # 2026-10-26 00:30+01:00 == 2026-10-25 23:30 UTC -> the next day
    order_at(event, "AFTER", dt.datetime(2026, 10, 25, 23, 30, tzinfo=dt.timezone.utc))

    reference = dt.datetime(2026, 10, 25, 12, 0, tzinfo=dt.timezone.utc)
    assert codes_placed_today(event, reference) == ["LASTX", "PASS1", "PASS2"]


@pytest.mark.django_db
def test_daily_windows_stay_gapless_and_disjoint_across_both_dst_changes(organizer):
    """Every order lands in exactly one day, on both switch days.

    The two tests above check the two switch days one at a time. This one checks
    the property that actually matters for a daily scheduled report: run it every
    day for four days across a switch and every order appears **exactly once**.
    A lost hour is a lost row, and nobody reconciles a daily export against the
    calendar.
    """
    event = berlin_event(organizer, slug="gapless")
    moments = {
        # around the spring change (2026-03-29)
        "SP001": dt.datetime(2026, 3, 27, 23, 30, tzinfo=dt.timezone.utc),
        "SP002": dt.datetime(2026, 3, 28, 22, 30, tzinfo=dt.timezone.utc),
        "SP003": dt.datetime(2026, 3, 28, 23, 30, tzinfo=dt.timezone.utc),
        "SP004": dt.datetime(2026, 3, 29, 21, 30, tzinfo=dt.timezone.utc),
        "SP005": dt.datetime(2026, 3, 29, 22, 30, tzinfo=dt.timezone.utc),
        "SP006": dt.datetime(2026, 3, 30, 21, 30, tzinfo=dt.timezone.utc),
    }
    for code, moment in moments.items():
        order_at(event, code, moment)

    seen: Dict[str, List[dt.date]] = {code: [] for code in moments}
    for day in range(27, 32):
        reference = dt.datetime(2026, 3, day, 12, 0, tzinfo=dt.timezone.utc)
        for code in codes_placed_today(event, reference):
            seen[code].append(dt.date(2026, 3, day))

    assert all(len(days) == 1 for days in seen.values()), seen
    assert seen["SP003"] == [dt.date(2026, 3, 29)], "00:30 CET on the switch day"
    assert seen["SP004"] == [dt.date(2026, 3, 29)], "23:30 CEST on the switch day"
    assert seen["SP005"] == [dt.date(2026, 3, 30)]


@pytest.mark.django_db
def test_relative_last_days_spans_the_switch_without_losing_a_day(organizer):
    """``relative_last_days: 3`` on 2026-10-26 = the 24th, 25th and 26th.

    The 25th is 25 hours long, so the window is 73 hours rather than 72. Counting
    in hours instead of calendar days would cut an hour off the far end and drop
    the order placed just after midnight on the 24th.
    """
    event = berlin_event(organizer, slug="lastdays")
    # 2026-10-24 00:30+02:00 == 2026-10-23 22:30 UTC -> inside (day 1 of 3)
    order_at(event, "DAY24", dt.datetime(2026, 10, 23, 22, 30, tzinfo=dt.timezone.utc))
    # 2026-10-23 23:30+02:00 == 2026-10-23 21:30 UTC -> outside, one day early
    order_at(event, "DAY23", dt.datetime(2026, 10, 23, 21, 30, tzinfo=dt.timezone.utc))
    # inside the repeated hour of the 25th
    order_at(event, "DAY25", dt.datetime(2026, 10, 25, 1, 30, tzinfo=dt.timezone.utc))
    # 2026-10-26 23:30+01:00 == 2026-10-26 22:30 UTC -> the last minutes of today
    order_at(event, "DAY26", dt.datetime(2026, 10, 26, 22, 30, tzinfo=dt.timezone.utc))

    document = definition(
        base="order",
        columns=columns("order.code"),
        filters={
            "op": "and",
            "children": [
                {
                    "field": "order.datetime",
                    "operator": "relative_last_days",
                    "value": 3,
                }
            ],
        },
    )
    reference = dt.datetime(2026, 10, 26, 12, 0, tzinfo=dt.timezone.utc)
    _, rows = run_report(document, event, now=reference)
    assert sorted(row[0] for row in rows) == ["DAY24", "DAY25", "DAY26"]


@pytest.mark.django_db
def test_a_scheduled_export_reevaluates_its_relative_filter_across_the_switch(
    registered, user_with_perms, organizer
):
    """The same schedule, run on two days that straddle the DST change.

    This is the whole reason relative operators exist. The filter is resolved
    when the export *runs*, so the file changes; if the window were frozen at
    save time the second run would mail the first run's rows.

    Runs through ``run_scheduled_exports``, the real ``periodic_task`` receiver,
    so ``EventTask`` (django-scopes), the owner's permission check and the mail
    are all part of the path.
    """
    from django.core import mail as djmail

    event = berlin_event(organizer, slug="sched")
    order_at(event, "DAY28", dt.datetime(2026, 3, 28, 12, 0, tzinfo=dt.timezone.utc))
    order_at(event, "DAY29", dt.datetime(2026, 3, 29, 12, 0, tzinfo=dt.timezone.utc))

    with scopes_disabled():
        ReportDefinition.objects.create(
            event=event,
            name="Today",
            identifier="today",
            definition=definition(
                base="order",
                columns=columns("order.code"),
                filters={
                    "op": "and",
                    "children": [
                        {"field": "order.datetime", "operator": "relative_today"}
                    ],
                },
            ),
        )
        schedule = ScheduledEventExport(event=event, owner=user_with_perms)
        schedule.export_identifier = exporters.CustomReportExporter.identifier
        schedule.export_form_data = {
            "_format": "default",
            contracts.EXPORT_FORM_REPORT_KEY: "today",
        }
        schedule.locale = "en"
        schedule.mail_subject = "Daily"
        schedule.mail_template = "see attachment"
        schedule.schedule_rrule = (
            "DTSTART:20260101T000000\nRRULE:FREQ=DAILY;INTERVAL=1;WKST=MO"
        )
        schedule.schedule_rrule_time = dt.time(9, 0)
        schedule.save()

    def run_on(day: int) -> str:
        djmail.outbox = []
        moment = dt.datetime(2026, 3, day, 10, 0, tzinfo=dt.timezone.utc)
        with scopes_disabled():
            ScheduledEventExport.objects.filter(pk=schedule.pk).update(
                schedule_next_run=moment - dt.timedelta(minutes=1), error_counter=0
            )
        with freeze_time(moment):
            run_scheduled_exports(None)
        assert len(djmail.outbox) == 1, djmail.outbox
        # Django hands text attachments back as ``str`` and binary ones as
        # ``bytes``; a CSV can arrive either way depending on the mime type.
        payload = djmail.outbox[0].attachments[0][1]
        if isinstance(payload, bytes):
            return payload.decode("utf-8-sig")
        return payload.lstrip("﻿")

    twenty_eighth = run_on(28)
    assert "DAY28" in twenty_eighth
    assert "DAY29" not in twenty_eighth

    twenty_ninth = run_on(29)
    assert "DAY29" in twenty_ninth
    assert "DAY28" not in twenty_ninth


@pytest.mark.django_db
def test_since_event_start_uses_the_events_own_start_instant(organizer):
    """``relative_since_event_start`` is an instant, not a midnight.

    The Berlin event starts 2026-06-03 09:00 UTC = 11:00 local. An order placed
    at 10:00 local on that day is *before* the start and must be out; the same
    report would include it if the operator rounded down to the start of the day.
    """
    event = berlin_event(organizer, slug="sincestart")
    order_at(event, "EARLY", dt.datetime(2026, 6, 3, 8, 0, tzinfo=dt.timezone.utc))
    order_at(event, "LATER", dt.datetime(2026, 6, 3, 10, 0, tzinfo=dt.timezone.utc))

    document = definition(
        base="order",
        columns=columns("order.code"),
        filters={
            "op": "and",
            "children": [
                {
                    "field": "order.datetime",
                    "operator": "relative_since_event_start",
                }
            ],
        },
    )
    reference = dt.datetime(2026, 7, 1, 12, 0, tzinfo=dt.timezone.utc)
    _, rows = run_report(document, event, now=reference)
    assert sorted(row[0] for row in rows) == ["LATER"]


# ===========================================================================
# 7. The signal wiring itself
# ===========================================================================
#
# Everything above takes the exporter registration for granted, because the
# ``registered`` fixture guarantees it. These two tests are about that
# guarantee: the first checks what production actually established, the second
# checks that this file gives it back.


@pytest.mark.django_db
def test_the_report_export_is_offered_exactly_once_on_both_levels(
    user_with_perms, main_event, organizer
):
    """The export appears once in the event *and* once in the organizer list.

    Deliberately without the ``registered`` fixture: the subject here is the
    wiring ``signals.py`` did at plugin import, and a fixture that guarantees
    the wiring cannot also be the witness for it.

    The interesting number is the ``1``. ``init_event_exporters`` instantiates
    one exporter per truthy signal response (services/export.py:198-222) and
    does not deduplicate, so every extra registration of
    :class:`CustomReportExporter` -- one receiver connected twice under two
    dispatch_uids, say, which is exactly what this module's old fixture did --
    becomes a second, identical entry in the export UI. Nothing raises,
    ``init_event_exporter`` keeps returning the first match, and the only
    visible symptom is a duplicated line on a page no test opens.
    """
    from pretix.base.services.export import (
        init_event_exporters,
        init_organizer_exporters,
    )

    identifier = exporters.CustomReportExporter.identifier
    with scope(organizer=organizer):
        event_level = [
            ex.identifier
            for ex in init_event_exporters(event=main_event, user=user_with_perms)
            if ex.identifier == identifier
        ]
        organizer_level = [
            ex.identifier
            for ex in init_organizer_exporters(
                organizer=organizer, user=user_with_perms
            )
            if ex.identifier == identifier
        ]

    assert event_level == [identifier], "event-level export registered != once"
    assert organizer_level == [identifier], "organizer-level export registered != once"

    # And it is ours, under the uids handoff/requests/exporter-dev-an-
    # integrator-signals.md asked for -- named literally, because other test
    # modules disconnect by them.
    assert WIRING_AT_IMPORT[DISPATCH_UID] is exporters.register_report_exporter
    assert (
        WIRING_AT_IMPORT[MULTI_DISPATCH_UID]
        is exporters.register_multievent_report_exporter
    )


# NOTE: keep this last. pytest runs a module's tests in definition order, so a
# check placed here has seen every ``registered`` teardown in this file.
def test_this_module_hands_the_signal_wiring_back_untouched(wiring_before_this_module):
    """The canary for the bug this file used to have.

    The first assertion is the one that belongs to this file: whatever we were
    handed, we hand back -- no receiver added, none removed, none rebound. The
    rest only runs when we really were handed the production wiring, so that
    this test keeps reporting *our* leaks and not another module's:
    ``tests/test_security.py`` still carries the old connect/disconnect pair
    (``security-reviewer``'s call), and it running first must not put a failure
    with the wrong name on it.
    """
    after = {
        dispatch_uid: connected_receiver(signal, dispatch_uid)
        for signal, _receiver, dispatch_uid in WIRING
    }
    assert after == wiring_before_this_module

    if wiring_before_this_module == WIRING_AT_IMPORT:
        # ``has_listeners`` and the identity check fail for different reasons:
        # a signal emptied wholesale versus our own uid dropped or rebound.
        assert register_data_exporters.has_listeners()
        assert register_multievent_data_exporters.has_listeners()
        for _signal, receiver, dispatch_uid in WIRING:
            assert after[dispatch_uid] is receiver, dispatch_uid
